from __future__ import annotations

import csv
import heapq
import json
import os
import re
import secrets
import shutil
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, time, timedelta, timezone
from dataclasses import replace
from functools import wraps
from io import BytesIO
from itertools import product
from pathlib import Path
from threading import Lock
from time import monotonic
from urllib.parse import urlsplit

import MetaTrader5 as mt5
import numpy as np
from flask import Flask, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash

from .app_paths import CACHE_DIR, DATA_DIR, INSTANCE_DIR, REPORT_DATA_FILE, RESOURCE_DIR, RESULTS_DIR, prepare_runtime
from .backtest_mt5 import BacktestConfig, TIMEFRAME_MINUTES, TIMEFRAMES, backtest, build_summary, clear_rates_cache, fetch_rates, parse_date, parse_time, write_outputs
from .market_data import (
    SOURCES,
    SOURCE_TIMEFRAMES,
    clear_market_data_cache,
    delta_history_status,
    delta_symbols,
    fetch_source_rates,
    normalize_source,
    validate_source_timeframe,
)
from .optimizer_engine import ScanContext, build_entry_layout, build_scan_context, evaluate_scan_batch, warm_optimizer_engine
from .algo_control import create_algo_blueprint


prepare_runtime()
app = Flask(__name__, static_folder=None, template_folder=str(RESOURCE_DIR), instance_path=str(INSTANCE_DIR))
Path(app.instance_path).mkdir(parents=True, exist_ok=True)
AUTH_FILE = Path(app.instance_path) / "auth.json"
SECRET_FILE = Path(app.instance_path) / "session_secret"
PROFILE_FILE = Path(app.instance_path) / "profile.json"
DEFAULT_USERNAME = "admin"
DEFAULT_PASSWORD_HASH = "scrypt:32768:8:1$QISOB6xfCKIB9Tef$1c28c2344d09c3fd43b055bc5d8699ce98e315cfa296c74193cd6ffe647c0d8a64bbef97ca68fbd6eb5f9b05e2927be83bd6f5b84dbb44cbc6242444e83a7a6d"
LOGIN_WINDOW_SECONDS = 300.0
LOGIN_MAX_ATTEMPTS = 5
_login_failures: dict[str, list[float]] = {}


def load_secret_key() -> str:
    configured = os.getenv("BACKTEST_SECRET_KEY")
    if configured:
        return configured
    if SECRET_FILE.exists():
        return SECRET_FILE.read_text(encoding="utf-8").strip()
    secret = secrets.token_urlsafe(48)
    SECRET_FILE.write_text(secret, encoding="utf-8")
    try:
        SECRET_FILE.chmod(0o600)
    except OSError:
        pass
    return secret


app.config.update(
    SECRET_KEY=load_secret_key(),
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Strict",
    SESSION_COOKIE_SECURE=os.getenv("BACKTEST_HTTPS", "").lower() in {"1", "true", "yes"},
    MAX_CONTENT_LENGTH=2 * 1024 * 1024,
)
HISTORY_CACHE_SECONDS = 300.0
_history_cache: dict[tuple[str, str, str, str, str], tuple[float, dict]] = {}
OPTIMIZER_MAX_COMBINATIONS = 2000000
OPTIMIZER_VISIBLE_RESULTS = 500
OPTIMIZER_CHECKPOINT_BATCH = 8192
OPTIMIZER_PROGRESS_EVERY = 10000
OPTIMIZER_RUNS_DIR = RESULTS_DIR / "optimizer_runs"
_optimizer_executor = ThreadPoolExecutor(max_workers=2, thread_name_prefix="optimizer")
_optimizer_warmup_executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="optimizer-warmup")
_optimizer_jobs: dict[str, dict] = {}
_optimizer_lock = Lock()
_saved_data_lock = Lock()
_optimizer_warmup_executor.submit(warm_optimizer_engine)
OPTIMIZER_ENTRY_PATTERNS = {
    "BOTH": 0,
    "BUY_ONLY": 1,
    "SELL_ONLY": -1,
}
OPTIMIZER_RESULT_SORTS = {"BALANCED", "WIN_RATE", "NET_POINTS", "LOWEST_TRADES", "WIN_POINTS"}
FALLBACK_MT5_SYMBOLS = ["BTCUSD#", "BTCUSD", "ETHUSD", "XAUUSD", "XAGUSD", "US30", "NAS100", "SPX500"]


def load_account() -> dict:
    if not AUTH_FILE.exists():
        provision_default_account()
    data = json.loads(AUTH_FILE.read_text(encoding="utf-8"))
    if not data.get("username") or not data.get("password_hash"):
        raise ValueError("Authentication configuration is invalid.")
    return data


def provision_default_account() -> None:
    if AUTH_FILE.exists():
        return
    payload = {
        "username": DEFAULT_USERNAME,
        "password_hash": DEFAULT_PASSWORD_HASH,
        "created_on": date.today().isoformat(),
    }
    temporary = AUTH_FILE.with_suffix(".tmp")
    temporary.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    temporary.replace(AUTH_FILE)
    try:
        AUTH_FILE.chmod(0o600)
    except OSError:
        pass


def load_profile() -> dict:
    if PROFILE_FILE.exists():
        data = json.loads(PROFILE_FILE.read_text(encoding="utf-8"))
    else:
        data = {}
    return {
        "name": str(data.get("name", session.get("username", ""))).strip(),
        "mobile_number": str(data.get("mobile_number", "")).strip(),
        "support_number": str(data.get("support_number", "")).strip(),
    }


def save_profile(data: dict) -> dict:
    profile = {
        "name": str(data.get("name", "")).strip(),
        "mobile_number": str(data.get("mobile_number", "")).strip(),
        "support_number": str(data.get("support_number", "")).strip(),
    }
    if not profile["name"]:
        raise ValueError("Name is required.")
    for key, label in (("mobile_number", "Mobile number"), ("support_number", "Customer support number")):
        value = profile[key]
        if len(value) > 30 or (value and not re.fullmatch(r"[0-9+() -]+", value)):
            raise ValueError(f"{label} is invalid.")
    temporary = PROFILE_FILE.with_suffix(".tmp")
    temporary.write_text(json.dumps(profile, indent=2), encoding="utf-8")
    temporary.replace(PROFILE_FILE)
    return profile


def change_password(data: dict) -> None:
    current_password = str(data.get("current_password", ""))
    new_password = str(data.get("new_password", ""))
    confirm_password = str(data.get("confirm_password", ""))
    account = load_account()
    if not check_password_hash(account["password_hash"], current_password):
        raise ValueError("Current password is incorrect.")
    if len(new_password) < 8:
        raise ValueError("New password must be at least 8 characters.")
    if new_password != confirm_password:
        raise ValueError("New passwords do not match.")
    if check_password_hash(account["password_hash"], new_password):
        raise ValueError("New password must be different from current password.")
    account["password_hash"] = generate_password_hash(new_password, method="scrypt")
    account["password_updated_on"] = date.today().isoformat()
    temporary = AUTH_FILE.with_suffix(".tmp")
    temporary.write_text(json.dumps(account, indent=2), encoding="utf-8")
    temporary.replace(AUTH_FILE)
    try:
        AUTH_FILE.chmod(0o600)
    except OSError:
        pass


provision_default_account()


@app.before_request
def cors_preflight():
    if request.method == "OPTIONS":
        return ("", 204)


def csrf_token() -> str:
    token = session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["csrf_token"] = token
    return token


def csrf_is_valid() -> bool:
    supplied = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token", "")
    return bool(supplied and secrets.compare_digest(supplied, session.get("csrf_token", "")))


def login_key() -> str:
    return request.remote_addr or "local"


def login_is_limited(key: str) -> bool:
    cutoff = monotonic() - LOGIN_WINDOW_SECONDS
    failures = [attempt for attempt in _login_failures.get(key, []) if attempt > cutoff]
    _login_failures[key] = failures
    return len(failures) >= LOGIN_MAX_ATTEMPTS


def record_login_failure(key: str) -> None:
    _login_failures.setdefault(key, []).append(monotonic())


def safe_redirect_target(target: str | None) -> str:
    if not target:
        return url_for("index")
    parsed = urlsplit(target)
    if parsed.netloc or parsed.scheme or not target.startswith("/") or target.startswith("//"):
        return url_for("index")
    return target


def is_authenticated() -> bool:
    return bool(session.get("username"))


def auth_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if is_authenticated():
            return view(*args, **kwargs)
        if request.path.startswith("/api/"):
            return jsonify({"error": "Authentication required."}), 401
        return redirect(url_for("login", next=request.full_path.rstrip("?")))

    return wrapped


app.register_blueprint(create_algo_blueprint(auth_required, csrf_is_valid))


def serialized_saved_data_write(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        with _saved_data_lock:
            return view(*args, **kwargs)

    return wrapped


def default_dates() -> tuple[str, str]:
    return "2025-06-01", "2026-05-25"


def load_latest_summary(source: str = "MT5") -> dict:
    source = normalize_source(source)
    source_path = RESULTS_DIR / source.lower() / "backtest_summary.json"
    path = source_path if source_path.exists() else RESULTS_DIR / "backtest_summary.json"
    if source != "MT5" and not source_path.exists():
        raise FileNotFoundError(f"No {source} backtest result found. Run a {source} backtest first.")
    if not path.exists():
        raise FileNotFoundError("No backtest result found. Run a backtest first.")
    return json.loads(path.read_text(encoding="utf-8"))


def clear_saved_data() -> None:
    for directory in (RESULTS_DIR, CACHE_DIR):
        if directory.exists():
            for item in directory.iterdir():
                if item.is_dir():
                    shutil.rmtree(item)
                else:
                    item.unlink()
        directory.mkdir(exist_ok=True)
    REPORT_DATA_FILE.write_text("window.BACKTEST_REPORT = null;\n", encoding="utf-8")


def clear_python_caches() -> None:
    skipped_directories = {".git", ".venv", "venv", "env", "node_modules"}
    cache_directories = {"__pycache__", ".pytest_cache", ".mypy_cache", ".ruff_cache"}
    compiled_suffixes = {".pyc", ".pyo"}
    for current_root, directories, files in os.walk(DATA_DIR, topdown=True):
        directories[:] = [name for name in directories if name not in skipped_directories]
        for name in tuple(directories):
            if name in cache_directories:
                shutil.rmtree(Path(current_root) / name)
                directories.remove(name)
        for name in files:
            if Path(name).suffix.lower() in compiled_suffixes:
                (Path(current_root) / name).unlink()


def add_sheet_rows(workbook: Workbook, title: str, rows: list[dict]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["No data"])
        return

    headers = list(rows[0].keys())
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")

    for row in rows:
        sheet.append([row.get(header) for header in headers])

    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = min(max(width, 12), 28)


def grouped_points(trades: list[dict], mode: str) -> list[dict]:
    groups: dict[str, dict] = {}
    for trade in trades:
        key = trade["trade_date"]
        if mode == "month":
            key = key[:7]
        elif mode == "year":
            key = key[:4]
        bucket = groups.setdefault(key, {"period": key, "trades": 0, "net_points": 0.0})
        bucket["trades"] += 1
        bucket["net_points"] += float(trade.get("pnl_points", 0))
    return [
        {"period": key, "trades": value["trades"], "net_points": round(value["net_points"], 2)}
        for key, value in sorted(groups.items())
    ]


def scan_values(data: dict, key: str, default: list, cast, maximum: int = 12) -> list:
    raw = data.get(key, default)
    values = raw if isinstance(raw, list) else str(raw).split(",")
    parsed = []
    for value in values:
        text = str(value).strip()
        if not text:
            continue
        item = cast(text)
        if item not in parsed:
            parsed.append(item)
    if not parsed:
        return list(default)
    if len(parsed) > maximum:
        raise ValueError(f"{key} accepts at most {maximum} values.")
    return parsed


def scan_entry_patterns(data: dict) -> list[str]:
    patterns = [value.upper() for value in scan_values(data, "entry_patterns", list(OPTIMIZER_ENTRY_PATTERNS), str, maximum=8)]
    invalid = [value for value in patterns if value not in OPTIMIZER_ENTRY_PATTERNS]
    if invalid:
        raise ValueError(f"Unsupported entry pattern: {', '.join(invalid)}")
    return patterns


def add_minutes(value, minutes: int):
    total = value.hour * 60 + value.minute + minutes
    total %= 24 * 60
    return time(total // 60, total % 60)


def minutes_of_day(value: time) -> int:
    return value.hour * 60 + value.minute


def ensure_time(value) -> time:
    if isinstance(value, time):
        return value
    return parse_time(str(value))


def minutes_after(anchor: int, value: time) -> int:
    value = ensure_time(value)
    minutes = minutes_of_day(value)
    while minutes < anchor:
        minutes += 24 * 60
    return minutes


def scan_time_profiles(data: dict) -> list[dict]:
    range_starts = scan_values(data, "range_start_values", [data.get("range_start", "08:30")], parse_time, maximum=24)
    durations = scan_values(data, "range_duration_values", [60], int, maximum=6)
    session_ends = scan_values(data, "session_end_values", [data.get("session_end", "19:30")], parse_time, maximum=24)
    entry_cutoff_values = scan_values(data, "entry_cutoff_values", [], parse_time, maximum=12) if str(data.get("entry_cutoff_values", "")).strip() else []
    profiles = []
    seen = set()
    rejected = {"range": 0, "session": 0, "entry": 0}
    first_hint = ""
    for range_start in range_starts:
        for duration in durations:
            if duration <= 0:
                raise ValueError("Range duration must be positive.")
            range_end = add_minutes(range_start, duration)
            session_start = range_end
            range_start_abs = minutes_of_day(range_start)
            range_end_abs = range_start_abs + duration
            session_start_abs = range_end_abs
            for session_end in session_ends:
                cutoffs_for_end = entry_cutoff_values or [session_end]
                for entry_cutoff in cutoffs_for_end:
                    raw_entry_cutoff_abs = minutes_after(range_start_abs, entry_cutoff)
                    session_end_abs = minutes_after(session_start_abs + 1, session_end)
                    if raw_entry_cutoff_abs < session_start_abs:
                        entry_cutoff_abs = session_start_abs
                        effective_entry_cutoff = session_start
                    else:
                        entry_cutoff_abs = minutes_after(session_start_abs, entry_cutoff)
                        effective_entry_cutoff = entry_cutoff
                    if range_end_abs <= range_start_abs:
                        rejected["range"] += 1
                        first_hint = first_hint or f"Range {range_start.strftime('%H:%M')} + {duration} min wraps past midnight; current optimizer supports same-day range windows."
                        continue
                    if session_end_abs <= session_start_abs:
                        rejected["session"] += 1
                        first_hint = first_hint or f"Range start {range_start.strftime('%H:%M')} makes session start {session_start.strftime('%H:%M')}; Force Exit must be after this time."
                        continue
                    if entry_cutoff_abs < session_start_abs or entry_cutoff_abs > session_end_abs:
                        rejected["entry"] += 1
                        first_hint = first_hint or f"Range start {range_start.strftime('%H:%M')} makes session start {session_start.strftime('%H:%M')}; Last Entry must be between {session_start.strftime('%H:%M')} and {session_end.strftime('%H:%M')}."
                        continue
                    key = (
                        range_start.strftime("%H:%M"),
                        range_end.strftime("%H:%M"),
                        session_start.strftime("%H:%M"),
                        effective_entry_cutoff.strftime("%H:%M"),
                        session_end.strftime("%H:%M"),
                    )
                    if key in seen:
                        continue
                    seen.add(key)
                    profiles.append(
                        {
                            "range_start": range_start,
                            "range_end": range_end,
                            "session_start": session_start,
                            "entry_cutoff": effective_entry_cutoff,
                            "session_end": session_end,
                        }
                    )
    if not profiles:
        detail = f" Rejected windows: range={rejected['range']}, session={rejected['session']}, entry={rejected['entry']}."
        hint = f" {first_hint}" if first_hint else ""
        raise ValueError(f"No valid optimizer time windows. Check range start, cutoff and force exit values.{detail}{hint}")
    return profiles


def optimizer_payload(data: dict) -> dict:
    source = normalize_source(data.get("data_source", "MT5"))
    symbol = str(data.get("symbol", "")).strip()
    if not symbol:
        raise ValueError("Symbol is required.")
    from_date = parse_date(data["from_date"])
    to_date = parse_date(data["to_date"])
    if from_date > to_date:
        raise ValueError("From date must be before to date.")
    available_timeframes = SOURCE_TIMEFRAMES[source]
    entry_timeframes = [value.upper() for value in scan_values(data, "entry_timeframes", available_timeframes, str)]
    trail_timeframes = [value.upper() for value in scan_values(data, "trail_timeframes", available_timeframes, str)]
    for timeframe in [*entry_timeframes, *trail_timeframes]:
        validate_source_timeframe(source, timeframe)
    entry_patterns = scan_entry_patterns(data)
    parameters = {
        "entry_buffer_pct": scan_values(data, "entry_buffer_values", [0.05, 0.20, 0.25], float),
        "stop_points": scan_values(data, "stop_points_values", [400.0, 500.0, 300.0, 200.0], float),
        "first_trail_profit": scan_values(data, "first_trail_profit_values", [400.0, 700.0, 600.0, 500.0], float),
        "first_trail_lock_loss": scan_values(data, "first_trail_lock_values", [300.0, 400.0, 200.0], float),
        "second_trail_profit": [700.0],
    }
    if any(value < 0 for values in parameters.values() for value in values):
        raise ValueError("Optimizer point and percentage settings must not be negative.")
    common = {
        "symbol": symbol,
        "from_date": from_date,
        "to_date": to_date,
        "data_source": source,
    }
    time_profiles = scan_time_profiles(data)
    total = (
        len(entry_timeframes)
        * len(trail_timeframes)
        * len(parameters["entry_buffer_pct"])
        * len(parameters["stop_points"])
        * len(parameters["first_trail_profit"])
        * len(parameters["first_trail_lock_loss"])
        * len(parameters["second_trail_profit"])
        * len(entry_patterns)
        * len(time_profiles)
    )
    maximum = min(max(int(data.get("max_combinations", OPTIMIZER_MAX_COMBINATIONS)), 1), OPTIMIZER_MAX_COMBINATIONS)
    if total > maximum:
        raise ValueError(f"Scan has {total} combinations. Reduce values or increase limit up to {OPTIMIZER_MAX_COMBINATIONS}.")
    return {
        "common": common,
        "entry_timeframes": entry_timeframes,
        "trail_timeframes": trail_timeframes,
        "entry_patterns": entry_patterns,
        "time_profiles": time_profiles,
        "parameters": parameters,
        "target_win_rate": min(max(float(data.get("target_win_rate", 70)), 0), 100),
        "minimum_trades": max(int(data.get("minimum_trades", 20)), 1),
        "result_sort": str(data.get("result_sort", "BALANCED")).upper() if str(data.get("result_sort", "BALANCED")).upper() in OPTIMIZER_RESULT_SORTS else "BALANCED",
        "stop_on_target": str(data.get("stop_on_target", "false")).lower() in {"true", "1", "on", "yes"},
        "total": total,
    }


def optimization_row(config: BacktestConfig, stats: dict, qualified: bool, tested: int, entry_pattern: str) -> dict:
    return {
        "rank": int(tested),
        "qualified": bool(qualified),
        "entry_pattern": entry_pattern,
        "timeframe": config.timeframe,
        "trail_timeframe": config.trail_timeframe,
        "range_start": config.range_start.strftime("%H:%M"),
        "range_end": config.range_end.strftime("%H:%M"),
        "session_start": config.session_start.strftime("%H:%M"),
        "entry_cutoff": config.entry_cutoff.strftime("%H:%M"),
        "session_end": config.session_end.strftime("%H:%M"),
        "entry_buffer_pct": round(float(config.entry_buffer_pct * 100), 6),
        "stop_points": float(config.stop_points),
        "first_trail_profit": float(config.first_trail_profit),
        "first_trail_lock_loss": float(config.first_trail_lock_loss),
        "second_trail_profit": float(config.second_trail_profit),
        "total_trades": int(stats["total_trades"]),
        "wins": int(stats["wins"]),
        "losses": int(stats["losses"]),
        "win_rate_pct": float(stats["win_rate_pct"]),
        "net_points": float(stats["net_points"]),
        "profit_factor": None if stats["profit_factor"] is None else float(stats["profit_factor"]),
        "max_drawdown_points": float(stats["max_drawdown_points"]),
    }


def optimizer_rank_key(row: dict, mode: str = "BALANCED") -> tuple:
    mode = mode if mode in OPTIMIZER_RESULT_SORTS else "BALANCED"
    qualified = bool(row.get("qualified"))
    win_rate = float(row.get("win_rate_pct") or 0)
    net_points = float(row.get("net_points") or 0)
    trades = int(row.get("total_trades") or 0)
    profit_factor = float(row.get("profit_factor") or 0)
    drawdown = abs(float(row.get("max_drawdown_points") or 0))
    if mode == "WIN_RATE":
        return (qualified, win_rate, net_points, profit_factor, trades)
    if mode == "NET_POINTS":
        return (qualified, net_points, win_rate, profit_factor, trades)
    if mode == "LOWEST_TRADES":
        return (qualified, -trades, win_rate, net_points, profit_factor)
    if mode == "WIN_POINTS":
        return (qualified, win_rate, net_points, -drawdown, trades)
    return (qualified, win_rate, net_points, trades, profit_factor, -drawdown)


def safe_optimizer_part(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value).strip().upper()) or "UNKNOWN"


def optimizer_scan_dir(source: str, symbol: str, scan_id: str) -> Path:
    return OPTIMIZER_RUNS_DIR / normalize_source(source).lower() / safe_optimizer_part(symbol) / safe_optimizer_part(scan_id)


def optimizer_output_path(source: str, symbol: str, scan_id: str) -> Path:
    return optimizer_scan_dir(source, symbol, scan_id) / "scan.json"


def optimizer_csv_path(source: str, symbol: str, scan_id: str) -> Path:
    return optimizer_scan_dir(source, symbol, scan_id) / "results.csv"


def optimizer_scan_config(payload: dict) -> dict:
    common = payload["common"]
    return {
        "symbol": common["symbol"],
        "data_source": common["data_source"],
        "from_date": common["from_date"].isoformat(),
        "to_date": common["to_date"].isoformat(),
        "time_profiles": [
            {
                "range_start": profile["range_start"].strftime("%H:%M"),
                "range_end": profile["range_end"].strftime("%H:%M"),
                "session_start": profile["session_start"].strftime("%H:%M"),
                "entry_cutoff": profile["entry_cutoff"].strftime("%H:%M"),
                "session_end": profile["session_end"].strftime("%H:%M"),
            }
            for profile in payload["time_profiles"]
        ],
        "entry_timeframes": payload["entry_timeframes"],
        "trail_timeframes": payload["trail_timeframes"],
        "entry_patterns": payload["entry_patterns"],
        "parameters": payload["parameters"],
        "target_win_rate": payload["target_win_rate"],
        "minimum_trades": payload["minimum_trades"],
        "result_sort": payload.get("result_sort", "BALANCED"),
        "stop_on_target": payload["stop_on_target"],
    }


def optimizer_public_result(job: dict, rows: list[dict]) -> dict:
    sort_mode = job.get("result_sort", "BALANCED")
    ranked = sorted(
        rows,
        key=lambda row: optimizer_rank_key(row, sort_mode),
        reverse=True,
    )
    for rank, row in enumerate(ranked, start=1):
        row["rank"] = rank
    result = dict(job)
    result["result_count"] = len(ranked)
    result["results_truncated"] = len(ranked) > OPTIMIZER_VISIBLE_RESULTS
    result["results"] = ranked[:OPTIMIZER_VISIBLE_RESULTS]
    return result


def save_optimizer_output(
    job: dict,
    payload: dict,
    rows: list[dict],
    csv_rows: list[dict] | None = None,
    replace_csv: bool = False,
    rows_are_ranked: bool = False,
    total_result_count: int | None = None,
) -> dict:
    common = payload["common"]
    if rows_are_ranked:
        stored = dict(job)
        stored["result_count"] = len(rows)
        stored["results_truncated"] = len(rows) > OPTIMIZER_VISIBLE_RESULTS
        stored["results"] = rows[:OPTIMIZER_VISIBLE_RESULTS]
    else:
        stored = optimizer_public_result(job, rows)
    if total_result_count is not None:
        stored["result_count"] = total_result_count
        stored["results_truncated"] = total_result_count > OPTIMIZER_VISIBLE_RESULTS
    stored["scan_config"] = optimizer_scan_config(payload)
    path = optimizer_output_path(common["data_source"], common["symbol"], job["scan_id"])
    path.parent.mkdir(parents=True, exist_ok=True)
    json_temp = path.with_suffix(".tmp")
    json_temp.write_text(json.dumps(stored, indent=2), encoding="utf-8")
    json_temp.replace(path)
    csv_path = optimizer_csv_path(common["data_source"], common["symbol"], job["scan_id"])
    if csv_rows:
        if replace_csv:
            csv_temp = csv_path.with_suffix(".tmp")
            handle = csv_temp.open("w", newline="", encoding="utf-8")
        else:
            csv_temp = None
            handle = csv_path.open("a", newline="", encoding="utf-8")
        with handle:
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
            if replace_csv or not csv_path.exists() or csv_path.stat().st_size == 0:
                writer.writeheader()
            writer.writerows(csv_rows)
        if csv_temp:
            csv_temp.replace(csv_path)
    return stored


def saved_optimizer_scans(source: str, symbol: str) -> list[dict]:
    directory = OPTIMIZER_RUNS_DIR / normalize_source(source).lower() / safe_optimizer_part(symbol)
    if not directory.exists():
        return []
    records = []
    for path in directory.glob("*/scan.json"):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            records.append(data)
        except (OSError, json.JSONDecodeError):
            continue
    records.sort(key=lambda item: item.get("created_at", ""), reverse=True)
    return records


def public_saved_optimizer_scan(row: dict) -> dict:
    active = row.get("scan_id") in _optimizer_jobs
    status = row.get("status")
    if status in {"queued", "running"} and not active:
        status = "interrupted"
    return {
        "scan_id": row["scan_id"],
        "status": status,
        "message": "Scan was interrupted. Saved partial results are available." if status == "interrupted" else row.get("message", ""),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "tested": row.get("tested", 0),
        "total": row.get("total", 0),
        "target_found": row.get("target_found", False),
        "best_win_rate": row.get("best_win_rate"),
        "scan_config": row.get("scan_config", {}),
    }


def execute_optimizer(job_id: str, payload: dict) -> None:
    started = monotonic()
    tested = 0
    results = []
    csv_buffer = []
    csv_started = False
    target_found = False
    frames: dict[tuple[str, bool], object] = {}
    layouts = {}
    contexts = {}
    common = payload["common"]
    debug_enabled = str(payload.get("debug", "")).lower() in {"1", "true", "yes", "on"}
    m5_derived_timeframes = {"M5", "M10", "M15", "M30", "H1"}
    use_m5_derived_frames = (
        common["data_source"] == "MT5"
        and set(payload["entry_timeframes"]).issubset(m5_derived_timeframes)
        and set(payload["trail_timeframes"]).issubset(m5_derived_timeframes)
    )
    has_overnight_windows = any(
        ensure_time(profile["session_end"]) <= ensure_time(profile["session_start"])
        or ensure_time(profile["entry_cutoff"]) < ensure_time(profile["session_start"])
        for profile in payload["time_profiles"]
    )

    def checkpoint(csv_rows: list[dict] | None = None, replace_csv: bool = False, **changes) -> None:
        def persist() -> None:
            changes["updated_at"] = datetime.now(timezone.utc).isoformat()
            with _optimizer_lock:
                _optimizer_jobs[job_id].update(changes)
                current_job = dict(_optimizer_jobs[job_id])
            stored = save_optimizer_output(
                current_job,
                payload,
                results,
                csv_rows=csv_rows,
                replace_csv=replace_csv,
                total_result_count=tested,
            )
            with _optimizer_lock:
                _optimizer_jobs[job_id].update(
                    {
                        "result_count": stored["result_count"],
                        "results_truncated": stored["results_truncated"],
                        "results": stored["results"],
                    }
                )

        if changes.get("status") in {"completed", "error"}:
            with _saved_data_lock:
                persist()
        else:
            persist()

    def update_progress(**changes) -> None:
        changes["updated_at"] = datetime.now(timezone.utc).isoformat()
        changes.setdefault("elapsed_seconds", round(monotonic() - started, 2))
        changes.setdefault("tested", tested)
        changes.setdefault("progress_pct", round((tested / payload["total"]) * 100, 1))
        if results:
            sort_mode = payload.get("result_sort", "BALANCED")
            key = lambda row: optimizer_rank_key(row, sort_mode)
            ranked = heapq.nlargest(
                OPTIMIZER_VISIBLE_RESULTS,
                results,
                key=key,
            )
            for rank, row in enumerate(ranked[:OPTIMIZER_VISIBLE_RESULTS], start=1):
                row["rank"] = rank
            changes["result_count"] = tested
            changes["results_truncated"] = tested > OPTIMIZER_VISIBLE_RESULTS
            changes["results"] = ranked
        with _optimizer_lock:
            _optimizer_jobs[job_id].update(changes)

    def debug_progress(message: str) -> None:
        if debug_enabled:
            update_progress(status="running", message=message)

    def resample_m5_frame(frame, timeframe: str):
        timeframe = timeframe.upper()
        if timeframe == "M5":
            return frame.copy()
        rule = f"{TIMEFRAME_MINUTES[timeframe]}min"
        indexed = frame.sort_values("time_ist").set_index("time_ist")
        resampled = indexed.resample(rule, origin="start_day", label="left", closed="left").agg(
            open=("open", "first"),
            high=("high", "max"),
            low=("low", "min"),
            close=("close", "last"),
            tick_volume=("tick_volume", "sum"),
        )
        resampled = resampled.dropna(subset=["open", "high", "low", "close"]).reset_index()
        resampled["trade_date"] = resampled["time_ist"].dt.date
        return resampled[["time_ist", "trade_date", "open", "high", "low", "close", "tick_volume"]].copy()

    def base_m5_frame(padded: bool):
        key = ("M5_BASE", True)
        if key not in frames:
            frame_config = BacktestConfig(
                symbol=common["symbol"],
                from_date=common["from_date"] - timedelta(days=1),
                to_date=common["to_date"] + timedelta(days=1) if has_overnight_windows else common["to_date"],
                data_source=common["data_source"],
                timeframe="M5",
            )
            debug_progress(f"DEBUG fetch MT5 M5 candles: {frame_config.from_date} to {frame_config.to_date}")
            frames[key] = fetch_source_rates(frame_config)
            debug_progress(f"DEBUG fetched M5 candles: {len(frames[key])} rows")
        if padded:
            return frames[key]
        return frames[key][frames[key]["trade_date"] >= common["from_date"]].copy()

    def candle_frame(timeframe: str, padded: bool):
        key = (timeframe, padded)
        if key not in frames:
            if use_m5_derived_frames:
                debug_progress(f"DEBUG build {timeframe}{' padded' if padded else ''} candles from M5 cache")
                frames[key] = resample_m5_frame(base_m5_frame(padded), timeframe)
                debug_progress(f"DEBUG ready {timeframe}{' padded' if padded else ''}: {len(frames[key])} rows")
            else:
                frame_config = BacktestConfig(
                    symbol=common["symbol"],
                    from_date=common["from_date"] - timedelta(days=1) if padded else common["from_date"],
                    to_date=common["to_date"] + timedelta(days=1) if has_overnight_windows else common["to_date"],
                    data_source=common["data_source"],
                    timeframe=timeframe,
                )
                debug_progress(f"DEBUG fetch {timeframe}{' padded' if padded else ''} candles from source")
                frames[key] = fetch_source_rates(frame_config)
                debug_progress(f"DEBUG ready {timeframe}{' padded' if padded else ''}: {len(frames[key])} rows")
        return frames[key]

    def time_profile_key(profile: dict) -> tuple[str, str, str, str, str]:
        return (
            profile["range_start"].strftime("%H:%M"),
            profile["range_end"].strftime("%H:%M"),
            profile["session_start"].strftime("%H:%M"),
            profile["entry_cutoff"].strftime("%H:%M"),
            profile["session_end"].strftime("%H:%M"),
        )

    def session_profile_key(profile: dict) -> tuple[str, str, str, str]:
        return (
            profile["range_start"].strftime("%H:%M"),
            profile["range_end"].strftime("%H:%M"),
            profile["session_start"].strftime("%H:%M"),
            profile["session_end"].strftime("%H:%M"),
        )

    def cutoff_ends_for_layout(layout, cutoff: time):
        cutoff_minute = cutoff.hour * 60 + cutoff.minute
        local_minutes = ((((layout.times_ns // 1_000_000_000) + 19800) % 86400) // 60).astype("int64")
        return np.asarray(
            [
                int(start + np.count_nonzero(local_minutes[start:end] <= cutoff_minute))
                for start, end in zip(layout.day_starts, layout.day_ends)
            ],
            dtype=np.int64,
        )

    def scan_context(entry_timeframe: str, trail_timeframe: str, profile: dict):
        profile["range_start"] = ensure_time(profile["range_start"])
        profile["range_end"] = ensure_time(profile["range_end"])
        profile["session_start"] = ensure_time(profile["session_start"])
        profile["entry_cutoff"] = ensure_time(profile["entry_cutoff"])
        profile["session_end"] = ensure_time(profile["session_end"])
        overnight_profile = profile["session_end"] <= profile["session_start"] or profile["entry_cutoff"] < profile["session_start"]
        base_profile_key = time_profile_key(profile) if overnight_profile else session_profile_key(profile)
        pair_key = (entry_timeframe, trail_timeframe, base_profile_key)
        if pair_key not in contexts:
            entry_df = candle_frame(entry_timeframe, False)
            layout_key = (entry_timeframe, base_profile_key)
            if layout_key not in layouts:
                layout_config = BacktestConfig(**common, **profile, timeframe=entry_timeframe, trail_timeframe=trail_timeframe)
                if len(layouts) % 25 == 0:
                    update_progress(
                        status="running",
                        message=f"Preparing layouts... {len(layouts)} built, {len(contexts)} contexts ready",
                    )
                layouts[layout_key] = build_entry_layout(entry_df, layout_config)
            trail_df = entry_df if trail_timeframe == entry_timeframe else candle_frame(trail_timeframe, True)
            if len(contexts) % 50 == 0:
                update_progress(
                    status="running",
                    message=f"Preparing trail contexts... {len(contexts)} ready",
                )
            contexts[pair_key] = build_scan_context(layouts[layout_key], trail_df, trail_timeframe)
        base_context = contexts[pair_key]
        if overnight_profile:
            return base_context
        adjusted_layout = replace(base_context.layout, cutoff_ends=cutoff_ends_for_layout(base_context.layout, profile["entry_cutoff"]))
        return ScanContext(layout=adjusted_layout, trail_lows=base_context.trail_lows, trail_highs=base_context.trail_highs)

    parameters = list(product(
        payload["parameters"]["entry_buffer_pct"],
        payload["parameters"]["stop_points"],
        payload["parameters"]["first_trail_profit"],
        payload["parameters"]["first_trail_lock_loss"],
        payload["parameters"]["second_trail_profit"],
    ))
    numeric_parameters = [
        (buffer_pct / 100, stop, first_profit, first_lock, second_profit)
        for buffer_pct, stop, first_profit, first_lock, second_profit in parameters
    ]
    try:
        update_progress(
            status="running",
            message="Loading M5 candles once and building selected scan timeframes..." if use_m5_derived_frames else "Loading candle data and scanning settings...",
        )
        stop_scan = False
        next_progress = OPTIMIZER_PROGRESS_EVERY
        last_progress = monotonic()
        for profile in payload["time_profiles"]:
            if stop_scan:
                break
            for entry_pattern in payload["entry_patterns"]:
                side_filter = OPTIMIZER_ENTRY_PATTERNS[entry_pattern]
                if stop_scan:
                    break
                for entry_tf in payload["entry_timeframes"]:
                    if stop_scan:
                        break
                    for trail_tf in payload["trail_timeframes"]:
                        if tested == 0:
                            update_progress(
                                tested=tested,
                                progress_pct=0.0,
                                elapsed_seconds=round(monotonic() - started, 2),
                                message=f"Preparing scan contexts for {entry_tf}/{trail_tf}...",
                            )
                        context = scan_context(entry_tf, trail_tf, profile)
                        for chunk_start in range(0, len(numeric_parameters), OPTIMIZER_CHECKPOINT_BATCH):
                            chunk_end = chunk_start + OPTIMIZER_CHECKPOINT_BATCH
                            chunk_parameters = parameters[chunk_start:chunk_end]
                            stats_rows = evaluate_scan_batch(context, numeric_parameters[chunk_start:chunk_end], side_filter=side_filter)
                            chunk_rows = []
                            for values, stats in zip(chunk_parameters, stats_rows):
                                tested += 1
                                buffer_pct, stop, first_profit, first_lock, second_profit = values
                                config = BacktestConfig(
                                    **common,
                                    **profile,
                                    timeframe=entry_tf,
                                    trail_timeframe=trail_tf,
                                    entry_buffer_pct=buffer_pct / 100,
                                    stop_points=stop,
                                    first_trail_profit=first_profit,
                                    first_trail_lock_loss=first_lock,
                                    second_trail_profit=second_profit,
                                )
                                qualified = stats["total_trades"] >= payload["minimum_trades"] and stats["win_rate_pct"] >= payload["target_win_rate"]
                                row = optimization_row(config, stats, qualified, tested, entry_pattern)
                                chunk_rows.append(row)
                                csv_buffer.append(row)
                                if qualified and payload["stop_on_target"]:
                                    stop_scan = True
                                    break
                            if chunk_rows:
                                results[:] = heapq.nlargest(
                                    OPTIMIZER_VISIBLE_RESULTS,
                                    [*results, *chunk_rows],
                                    key=lambda row: optimizer_rank_key(row, payload.get("result_sort", "BALANCED")),
                                )
                            target_found = target_found or any(row["qualified"] for row in chunk_rows)
                            now = monotonic()
                            report_progress = (
                                stop_scan
                                or tested >= payload["total"]
                                or tested >= next_progress
                                or now - last_progress >= 1.0
                            )
                            if report_progress:
                                update_progress(
                                    tested=tested,
                                    progress_pct=round((tested / payload["total"]) * 100, 1),
                                    best_win_rate=max(row["win_rate_pct"] for row in results),
                                    elapsed_seconds=round(monotonic() - started, 2),
                                    message=f"Tested {tested} of {payload['total']} settings",
                                )
                                last_progress = now
                            if stop_scan or tested >= payload["total"] or tested >= next_progress:
                                checkpoint(
                                    csv_rows=csv_buffer,
                                    replace_csv=not csv_started,
                                    status="running",
                                    tested=tested,
                                    progress_pct=round((tested / payload["total"]) * 100, 1),
                                    best_win_rate=max(row["win_rate_pct"] for row in results),
                                    elapsed_seconds=round(monotonic() - started, 2),
                                    message=f"Tested {tested} of {payload['total']} settings",
                                )
                                csv_buffer.clear()
                                csv_started = True
                                while next_progress <= tested:
                                    next_progress += OPTIMIZER_PROGRESS_EVERY
                            if stop_scan:
                                break
                        if stop_scan:
                            break
            layouts.clear()
            contexts.clear()
        found = target_found
        final = {
            "status": "completed",
            "message": "Target reached." if found else "Scan completed without reaching target.",
            "tested": tested,
            "total": payload["total"],
            "progress_pct": 100.0 if found and payload["stop_on_target"] else round((tested / payload["total"]) * 100, 1),
            "stopped_early": found and payload["stop_on_target"] and tested < payload["total"],
            "target_found": found,
            "target_win_rate": payload["target_win_rate"],
            "minimum_trades": payload["minimum_trades"],
            "elapsed_seconds": round(monotonic() - started, 2),
        }
        ranked_rows = sorted(
            results,
            key=lambda row: optimizer_rank_key(row, payload.get("result_sort", "BALANCED")),
            reverse=True,
        )
        for rank, row in enumerate(ranked_rows, start=1):
            row["rank"] = rank
        results[:] = ranked_rows
        checkpoint(csv_rows=csv_buffer, replace_csv=not csv_started, **final)
    except Exception as exc:
        checkpoint(
            csv_rows=csv_buffer,
            replace_csv=not csv_started,
            status="error",
            message=str(exc),
            tested=tested,
            progress_pct=round((tested / payload["total"]) * 100, 1),
            elapsed_seconds=round(monotonic() - started, 2),
        )


def mark_optimizer_submit_error(job_id: str, future) -> None:
    try:
        future.result()
    except Exception as exc:
        with _optimizer_lock:
            job = _optimizer_jobs.get(job_id)
            if job and job.get("status") in {"queued", "running"}:
                job.update(
                    {
                        "status": "error",
                        "message": str(exc),
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }
                )


@app.after_request
def add_headers(response):
    response.headers["Cache-Control"] = "no-store"
    response.headers["Referrer-Policy"] = "same-origin"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    origin = request.headers.get("Origin")
    if origin in {"null", "http://127.0.0.1:5000", "http://localhost:5000"}:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-CSRF-Token"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response


@app.route("/login", methods=["GET", "POST"])
def login():
    if is_authenticated():
        return redirect(url_for("index"))
    error = None
    target = request.form.get("next") or request.args.get("next", "")
    if request.method == "POST":
        key = login_key()
        if not csrf_is_valid():
            error = "This form expired. Please try again."
        elif login_is_limited(key):
            error = "Too many failed attempts. Wait five minutes before trying again."
        else:
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            try:
                account = load_account()
            except (OSError, ValueError, json.JSONDecodeError):
                return render_template("login.html", csrf_token=csrf_token(), error="Account configuration could not be read.", next=target), 500
            valid_username = secrets.compare_digest(username, account["username"])
            valid_password = check_password_hash(account["password_hash"], password)
            if valid_username and valid_password:
                _login_failures.pop(key, None)
                session.clear()
                session.permanent = True
                session["username"] = account["username"]
                csrf_token()
                return redirect(safe_redirect_target(target))
            record_login_failure(key)
            error = "Invalid user ID or password."
    return render_template("login.html", csrf_token=csrf_token(), error=error, next=target)


@app.post("/logout")
@auth_required
def logout():
    if not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    session.clear()
    return redirect(url_for("login"))


@app.get("/styles.css")
def public_styles():
    return send_from_directory(RESOURCE_DIR, "styles.css")


@app.get("/vendor/highcharts/<path:filename>")
@auth_required
def highcharts_asset(filename: str):
    if filename != "highstock.js":
        return jsonify({"error": "Asset not found."}), 404
    return send_from_directory(RESOURCE_DIR / "vendor" / "highcharts", filename)


@app.get("/")
@auth_required
def index():
    return send_from_directory(RESOURCE_DIR, "index.html")


@app.get("/profile")
@auth_required
def profile_page():
    return send_from_directory(RESOURCE_DIR, "profile.html")


@app.get("/api/health")
def health():
    return jsonify({"ok": True, "server": "mt5-backtest", "version": 3})


@app.get("/api/defaults")
@auth_required
def defaults():
    start, end = default_dates()
    return jsonify(
        {
            "from_date": start,
            "to_date": end,
            "sources": SOURCES,
            "source_timeframes": SOURCE_TIMEFRAMES,
            "defaults": {
                "data_source": "MT5",
                "timeframe": "M5",
                "trail_timeframe": "M15",
                "symbol": "BTCUSD",
                "entry_pattern": "BOTH",
                "range_start": "08:30",
                "range_end": "09:30",
                "session_start": "09:30",
                "entry_cutoff": "18:00",
                "session_end": "19:30",
                "entry_buffer_pct": 0.25,
                "stop_points": 500,
                "first_trail_profit": 400,
                "first_trail_lock_loss": 200,
                "second_trail_profit": 700,
            },
        }
    )


@app.get("/api/symbols")
@auth_required
def symbols():
    try:
        source = normalize_source(request.args.get("source", "MT5"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    if source == "DELTA":
        try:
            return jsonify(delta_symbols())
        except Exception as exc:
            return jsonify({"error": f"Delta symbol list failed: {exc}"}), 400

    if not mt5.initialize():
        return jsonify(FALLBACK_MT5_SYMBOLS)

    try:
        all_symbols = mt5.symbols_get()
    finally:
        mt5.shutdown()

    if not all_symbols:
        return jsonify(FALLBACK_MT5_SYMBOLS)

    priority = ("BTC", "ETH", "XRP", "SOL", "DOGE", "BNB", "ADA", "XAU", "XAG", "US30", "NAS", "SPX")
    names = sorted({symbol.name for symbol in all_symbols} | set(FALLBACK_MT5_SYMBOLS))
    names.sort(key=lambda name: (not any(term in name.upper() for term in priority), name.upper()))
    return jsonify(names)


@app.get("/api/history-range")
@auth_required
def history_range():
    try:
        source = normalize_source(request.args.get("source", "MT5"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    symbol = request.args.get("symbol", "BTCUSD").strip()
    timeframe = request.args.get("timeframe", "M5").upper()

    default_from, default_to = default_dates()
    requested_from = request.args.get("from_date", default_from)
    requested_to = request.args.get("to_date", default_to)
    try:
        validate_source_timeframe(source, timeframe)
        from_date = parse_date(requested_from)
        to_date = parse_date(requested_to)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    cache_key = (source, symbol, timeframe, requested_from, requested_to)
    cached = _history_cache.get(cache_key)
    if cached and monotonic() - cached[0] < HISTORY_CACHE_SECONDS:
        return jsonify(cached[1])

    if source == "DELTA":
        try:
            result = delta_history_status(symbol, timeframe, from_date, to_date)
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400
        _history_cache[cache_key] = (monotonic(), result)
        return jsonify(result)

    try:
        config = BacktestConfig(
            symbol=symbol,
            from_date=from_date,
            to_date=to_date,
            data_source=source,
            timeframe=timeframe,
        )
        df = fetch_rates(config)
    except Exception as exc:
        if "No MT5 candle data returned" not in str(exc):
            return jsonify({"error": str(exc)}), 400
        result = {"symbol": symbol, "timeframe": timeframe, "available": False}
    else:
        if df.empty:
            result = {"symbol": symbol, "timeframe": timeframe, "available": False}
        else:
            result = {
                "symbol": symbol,
                "timeframe": timeframe,
                "available": True,
                "from_date": min(df["trade_date"]).isoformat(),
                "to_date": max(df["trade_date"]).isoformat(),
                "from_time": df.iloc[0]["time_ist"].isoformat(),
                "to_time": df.iloc[-1]["time_ist"].isoformat(),
            }
    if result["available"]:
        _history_cache[cache_key] = (monotonic(), result)
    return jsonify(result)


@app.get("/api/export/excel")
@auth_required
def export_excel():
    try:
        summary = load_latest_summary(request.args.get("source", "MT5"))
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        config_rows = [{"field": key, "value": value} for key, value in summary.get("config", {}).items()]
        stats_rows = [{"metric": key, "value": value} for key, value in summary.get("stats", {}).items()]
        trades = summary.get("trades", [])

        add_sheet_rows(workbook, "Config", config_rows)
        add_sheet_rows(workbook, "Stats", stats_rows)
        add_sheet_rows(workbook, "Trades", trades)
        add_sheet_rows(workbook, "Daily Points", grouped_points(trades, "day"))
        add_sheet_rows(workbook, "Monthly Points", grouped_points(trades, "month"))
        add_sheet_rows(workbook, "Yearly Points", grouped_points(trades, "year"))

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        symbol = summary.get("config", {}).get("symbol", "backtest")
        filename = f"{symbol}_backtest_export.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/data/clear")
@auth_required
def clear_all_data():
    if not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    with _saved_data_lock:
        with _optimizer_lock:
            active_job = next(
                (job for job in _optimizer_jobs.values() if job.get("status") in {"queued", "running"}),
                None,
            )
            if active_job:
                return jsonify({"error": "An optimizer scan is running. Wait for it to finish before clearing all data."}), 409
            try:
                clear_saved_data()
                clear_python_caches()
                _optimizer_jobs.clear()
                _history_cache.clear()
                clear_rates_cache()
                clear_market_data_cache()
            except OSError as exc:
                return jsonify({"error": f"Could not clear saved data: {exc}"}), 500
    return jsonify({"message": "All saved results, candle data, and Python caches have been cleared."})


@app.post("/api/backtest")
@auth_required
@serialized_saved_data_write
def run_backtest():
    if not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    data = request.get_json(force=True)
    try:
        source = normalize_source(data.get("data_source", "MT5"))
        config = BacktestConfig(
            symbol=str(data.get("symbol", "")).strip(),
            from_date=parse_date(data["from_date"]),
            to_date=parse_date(data["to_date"]),
            data_source=source,
            timeframe=str(data.get("timeframe", "M5")).upper(),
            trail_timeframe=str(data.get("trail_timeframe", data.get("timeframe", "M5"))).upper(),
            entry_pattern=str(data.get("entry_pattern", "BOTH")).upper(),
            range_start=parse_time(data.get("range_start", "08:30")),
            range_end=parse_time(data.get("range_end", "09:30")),
            session_start=parse_time(data.get("session_start", "09:30")),
            entry_cutoff=parse_time(data.get("entry_cutoff", "18:00")),
            session_end=parse_time(data.get("session_end", "19:30")),
            entry_buffer_pct=float(data.get("entry_buffer_pct", 0.25)) / 100,
            stop_points=float(data.get("stop_points", 500)),
            first_trail_profit=float(data.get("first_trail_profit", 400)),
            first_trail_lock_loss=float(data.get("first_trail_lock_loss", 200)),
            second_trail_profit=float(data.get("second_trail_profit", 700)),
        )
        if not config.symbol:
            raise ValueError("Symbol is required.")
        validate_source_timeframe(source, config.timeframe)
        validate_source_timeframe(source, config.trail_timeframe)
        if config.entry_pattern not in OPTIMIZER_ENTRY_PATTERNS:
            raise ValueError("Unsupported entry pattern.")
        if config.range_start >= config.range_end:
            raise ValueError("Range start must be before range end.")
        if config.session_start >= config.session_end:
            raise ValueError("Session start must be before session end.")
        if config.entry_cutoff < config.session_start or config.entry_cutoff > config.session_end:
            raise ValueError("Last entry time must be within the trading session.")

        df = fetch_source_rates(config)
        if not df.empty:
            config.data_from_date = min(df["trade_date"])
            config.data_to_date = max(df["trade_date"])
        if config.trail_timeframe == config.timeframe:
            trail_df = df
        else:
            trail_config = BacktestConfig(
                symbol=config.symbol,
                from_date=config.from_date - timedelta(days=1),
                to_date=config.to_date,
                data_source=config.data_source,
                timeframe=config.trail_timeframe,
            )
            trail_df = fetch_source_rates(trail_config)
        trades = backtest(df, config, trail_df)
        summary = build_summary(trades, config)
        write_outputs(summary)
        return jsonify(summary)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/optimizer/start")
@auth_required
def start_optimizer():
    if not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    try:
        payload = optimizer_payload(request.get_json(force=True))
        common = payload["common"]
        with _optimizer_lock:
            active_job = next(
                (
                    job
                    for job in _optimizer_jobs.values()
                    if job.get("status") in {"queued", "running"}
                    and job.get("source") == common["data_source"]
                    and job.get("symbol") == common["symbol"]
                ),
                None,
            )
        if active_job:
            return jsonify(
                {
                    "job_id": active_job["scan_id"],
                    "scan_id": active_job["scan_id"],
                    "source": common["data_source"],
                    "symbol": common["symbol"],
                    "total": active_job["total"],
                    "reused": True,
                }
            )
        job_id = f"{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}_{secrets.token_hex(4)}"
        created_at = datetime.now(timezone.utc).isoformat()
        job = {
            "scan_id": job_id,
            "source": common["data_source"],
            "symbol": common["symbol"],
            "status": "queued",
            "message": "Scan queued...",
            "tested": 0,
            "total": payload["total"],
            "progress_pct": 0.0,
            "target_win_rate": payload["target_win_rate"],
            "minimum_trades": payload["minimum_trades"],
            "result_sort": payload["result_sort"],
            "created_at": created_at,
            "updated_at": created_at,
            "result_count": 0,
            "results_truncated": False,
            "results": [],
        }
        with _optimizer_lock:
            _optimizer_jobs[job_id] = job
        save_optimizer_output(job, payload, [])
        future = _optimizer_executor.submit(execute_optimizer, job_id, payload)
        future.add_done_callback(lambda item, scan_id=job_id: mark_optimizer_submit_error(scan_id, item))
        return jsonify(
            {
                "job_id": job_id,
                "scan_id": job_id,
                "source": common["data_source"],
                "symbol": common["symbol"],
                "total": payload["total"],
            }
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/optimizer/jobs/<job_id>")
@auth_required
def optimizer_job(job_id: str):
    with _optimizer_lock:
        job = _optimizer_jobs.get(job_id)
        result = dict(job) if job else None
    if result is None:
        return jsonify({"error": "Optimizer scan not found."}), 404
    return jsonify(result)


@app.get("/api/optimizer/scans")
@auth_required
def optimizer_scans():
    try:
        source = normalize_source(request.args.get("source", "MT5"))
        symbol = request.args.get("symbol", "").strip()
        if not symbol:
            raise ValueError("Symbol is required.")
        records = saved_optimizer_scans(source, symbol)
        summary = [public_saved_optimizer_scan(row) for row in records[:50]]
        return jsonify(summary)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/optimizer/scans/<scan_id>")
@auth_required
def saved_optimizer_scan(scan_id: str):
    try:
        source = normalize_source(request.args.get("source", "MT5"))
        symbol = request.args.get("symbol", "").strip()
        if not symbol:
            raise ValueError("Symbol is required.")
        path = optimizer_output_path(source, symbol, scan_id)
        if not path.exists():
            return jsonify({"error": "Optimizer scan not found."}), 404
        result = json.loads(path.read_text(encoding="utf-8"))
        with _optimizer_lock:
            active = scan_id in _optimizer_jobs
        if result.get("status") in {"queued", "running"} and not active:
            result["status"] = "interrupted"
            result["message"] = "Scan was interrupted. Saved partial results are available."
        return jsonify(result)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/optimizer/latest")
@auth_required
def latest_optimizer():
    try:
        source = normalize_source(request.args.get("source", "MT5"))
        symbol = request.args.get("symbol", "").strip()
        if not symbol:
            raise ValueError("Symbol is required.")
        scans = saved_optimizer_scans(source, symbol)
        if not scans:
            return jsonify({"error": "No optimizer scan result found."}), 404
        return jsonify(scans[0])
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/optimizer/export")
@auth_required
def export_optimizer():
    try:
        source = normalize_source(request.args.get("source", "MT5"))
        symbol = request.args.get("symbol", "").strip()
        if not symbol:
            raise ValueError("Symbol is required.")
        scan_id = request.args.get("scan_id", "").strip()
        if not scan_id:
            scans = saved_optimizer_scans(source, symbol)
            if not scans:
                return jsonify({"error": "No optimizer scan result found."}), 404
            scan_id = scans[0]["scan_id"]
        path = optimizer_csv_path(source, symbol, scan_id)
        if not path.exists():
            return jsonify({"error": "No optimizer scan result found."}), 404
        filename = f"{safe_optimizer_part(symbol).lower()}_{source.lower()}_{scan_id}_optimizer.csv"
        return send_file(path, as_attachment=True, download_name=filename)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/candles", methods=["GET", "POST", "OPTIONS"])
@auth_required
def candles():
    if request.method == "OPTIONS":
        return jsonify({"ok": True})
    if request.method == "POST" and not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    data = request.args.to_dict() if request.method == "GET" else request.get_json(force=True)
    try:
        source = normalize_source(data.get("data_source", "MT5"))
        from_date = parse_date(data.get("from_date") or data["trade_date"])
        to_date = parse_date(data.get("to_date") or data.get("trade_date") or data["from_date"])
        config = BacktestConfig(
            symbol=str(data.get("symbol", "")).strip(),
            from_date=from_date,
            to_date=to_date,
            data_source=source,
            timeframe=str(data.get("timeframe", "M5")).upper(),
        )
        if not config.symbol:
            raise ValueError("Symbol is required.")
        validate_source_timeframe(source, config.timeframe)
        df = fetch_source_rates(config)
        rows = []
        for _, row in df.iterrows():
            rows.append(
                {
                    "time": int(row["time_ist"].timestamp()),
                    "open": round(float(row["open"]), 2),
                    "high": round(float(row["high"]), 2),
                    "low": round(float(row["low"]), 2),
                    "close": round(float(row["close"]), 2),
                }
            )
        return jsonify({"candles": rows, "timezone": "Asia/Kolkata"})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/session")
@auth_required
def current_session():
    return jsonify({"username": session["username"], "csrf_token": csrf_token()})


@app.route("/api/profile", methods=["GET", "POST"])
@auth_required
def profile_details():
    if request.method == "POST" and not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    try:
        if request.method == "POST":
            return jsonify(save_profile(request.get_json(force=True)))
        return jsonify(load_profile())
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/profile/password")
@auth_required
def update_profile_password():
    if not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    try:
        change_password(request.get_json(force=True))
        return jsonify({"message": "Password updated successfully."})
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/report_data.js")
@auth_required
def latest_report_javascript():
    path = REPORT_DATA_FILE
    if path.is_file():
        return send_from_directory(path.resolve().parent, path.name)
    return app.response_class("window.BACKTEST_REPORT = null;\n", mimetype="application/javascript")


@app.get("/<path:path>")
@auth_required
def static_files(path: str):
    return send_from_directory(RESOURCE_DIR, "index.html")


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
