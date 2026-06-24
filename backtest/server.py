from __future__ import annotations

import csv
import heapq
import json
import os
import re
import secrets
import shutil
import sqlite3
import sys
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, time, timedelta, timezone
from functools import wraps
from io import BytesIO, StringIO
from itertools import product
from pathlib import Path
from threading import Lock
from time import monotonic
from urllib.parse import urlsplit

import MetaTrader5 as mt5
import numpy as np
from flask import Flask, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    __package__ = "backtest"

from .app_paths import CACHE_DIR, DATA_DIR, INSTANCE_DIR, REPORT_DATA_FILE, RESOURCE_DIR, RESULTS_DIR, prepare_runtime
from .backtest_mt5 import (
    BacktestConfig,
    TIMEFRAME_MINUTES,
    TIMEFRAMES,
    backtest,
    build_summary,
    clear_rates_cache,
    fetch_rates,
    normalize_distance_unit,
    parse_date,
    parse_time,
    write_outputs,
)
from .market_data import (
    SOURCES,
    SOURCE_TIMEFRAMES,
    clear_market_data_cache,
    delta_history_status,
    delta_symbols,
    fetch_source_rates,
    normalize_source,
    validate_source_timeframe,
)
from .optimizer_engine import build_entry_layout, build_execution_layout, build_scan_context, evaluate_scan_batch_metrics, warm_optimizer_engine


prepare_runtime()
app = Flask(__name__, static_folder=None, template_folder=str(RESOURCE_DIR), instance_path=str(INSTANCE_DIR))
Path(app.instance_path).mkdir(parents=True, exist_ok=True)
AUTH_FILE = Path(app.instance_path) / "auth.json"
SECRET_FILE = Path(app.instance_path) / "session_secret"
PROFILE_FILE = Path(app.instance_path) / "profile.json"
DEFAULT_USERNAME = "admin"
DEFAULT_PASSWORD_HASH = "scrypt:32768:8:1$QISOB6xfCKIB9Tef$1c28c2344d09c3fd43b055bc5d8699ce98e315cfa296c74193cd6ffe647c0d8a64bbef97ca68fbd6eb5f9b05e2927be83bd6f5b84dbb44cbc6242444e83a7a6d"
LOGIN_WINDOW_SECONDS = 300.0
LOGIN_MAX_ATTEMPTS = 5
_login_failures: dict[str, list[float]] = {}


def load_secret_key() -> str:
    configured = os.getenv("BACKTEST_SECRET_KEY")
    if configured:
        return configured
    if SECRET_FILE.exists():
        return SECRET_FILE.read_text(encoding="utf-8").strip()
    secret = secrets.token_urlsafe(48)
    SECRET_FILE.write_text(secret, encoding="utf-8")
    try:
        SECRET_FILE.chmod(0o600)
    except OSError:
        pass
    return secret


app.config.update(
    SECRET_KEY=load_secret_key(),
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Strict",
    SESSION_COOKIE_SECURE=os.getenv("BACKTEST_HTTPS", "").lower() in {"1", "true", "yes"},
    MAX_CONTENT_LENGTH=2 * 1024 * 1024,
)
HISTORY_CACHE_SECONDS = 300.0
_history_cache: dict[tuple[str, str, str, str, str], tuple[float, dict]] = {}


def optimizer_int_env(name: str, default: int, minimum: int) -> int:
    try:
        return max(int(os.getenv(name, str(default))), minimum)
    except (TypeError, ValueError):
        return default


OPTIMIZER_MAX_COMBINATIONS = 10_000_000
OPTIMIZER_VISIBLE_RESULTS = 500
OPTIMIZER_CHECKPOINT_BATCH = optimizer_int_env("OPTIMIZER_CHECKPOINT_BATCH", 65536, 1024)
OPTIMIZER_PROGRESS_EVERY = optimizer_int_env("OPTIMIZER_PROGRESS_EVERY", 25000, 1000)
OPTIMIZER_CHECKPOINT_EVERY = max(optimizer_int_env("OPTIMIZER_CHECKPOINT_EVERY", 250000, 1000), OPTIMIZER_PROGRESS_EVERY)
OPTIMIZER_SAVE_FULL_CSV_DEFAULT = os.getenv("OPTIMIZER_SAVE_FULL_CSV", "").lower() in {"1", "true", "yes", "on"}
OPTIMIZER_RUNS_DIR = RESULTS_DIR / "optimizer_runs"
OPTIMIZER_DB_FILE = OPTIMIZER_RUNS_DIR / "optimizer.sqlite3"
OPTIMIZER_ENGINE_VERSION = 3
_optimizer_executor = ThreadPoolExecutor(max_workers=2, thread_name_prefix="optimizer")
_optimizer_warmup_executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="optimizer-warmup")
_optimizer_jobs: dict[str, dict] = {}
_optimizer_lock = Lock()
_optimizer_db_lock = Lock()
_saved_data_lock = Lock()
_optimizer_warmup_executor.submit(warm_optimizer_engine)
OPTIMIZER_ENTRY_PATTERNS = {
    "BOTH": 0,
    "BUY_ONLY": 1,
    "SELL_ONLY": -1,
}
OPTIMIZER_RESULT_SORTS = {"BALANCED", "WIN_RATE", "NET_POINTS", "LOWEST_TRADES", "WIN_POINTS"}
FALLBACK_MT5_SYMBOLS = ["BTCUSD#", "BTCUSD", "ETHUSD", "XAUUSD", "XAGUSD", "US30", "NAS100", "SPX500"]


def load_account() -> dict:
    if not AUTH_FILE.exists():
        provision_default_account()
    data = json.loads(AUTH_FILE.read_text(encoding="utf-8"))
    if not data.get("username") or not data.get("password_hash"):
        raise ValueError("Authentication configuration is invalid.")
    return data


def provision_default_account() -> None:
    if AUTH_FILE.exists():
        return
    payload = {
        "username": DEFAULT_USERNAME,
        "password_hash": DEFAULT_PASSWORD_HASH,
        "created_on": date.today().isoformat(),
    }
    temporary = AUTH_FILE.with_suffix(".tmp")
    temporary.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    temporary.replace(AUTH_FILE)
    try:
        AUTH_FILE.chmod(0o600)
    except OSError:
        pass


def load_profile() -> dict:
    if PROFILE_FILE.exists():
        data = json.loads(PROFILE_FILE.read_text(encoding="utf-8"))
    else:
        data = {}
    return {
        "name": str(data.get("name", session.get("username", ""))).strip(),
        "mobile_number": str(data.get("mobile_number", "")).strip(),
        "support_number": str(data.get("support_number", "")).strip(),
    }


def save_profile(data: dict) -> dict:
    profile = {
        "name": str(data.get("name", "")).strip(),
        "mobile_number": str(data.get("mobile_number", "")).strip(),
        "support_number": str(data.get("support_number", "")).strip(),
    }
    if not profile["name"]:
        raise ValueError("Name is required.")
    for key, label in (("mobile_number", "Mobile number"), ("support_number", "Customer support number")):
        value = profile[key]
        if len(value) > 30 or (value and not re.fullmatch(r"[0-9+() -]+", value)):
            raise ValueError(f"{label} is invalid.")
    temporary = PROFILE_FILE.with_suffix(".tmp")
    temporary.write_text(json.dumps(profile, indent=2), encoding="utf-8")
    temporary.replace(PROFILE_FILE)
    return profile


def change_password(data: dict) -> None:
    current_password = str(data.get("current_password", ""))
    new_password = str(data.get("new_password", ""))
    confirm_password = str(data.get("confirm_password", ""))
    account = load_account()
    if not check_password_hash(account["password_hash"], current_password):
        raise ValueError("Current password is incorrect.")
    if len(new_password) < 8:
        raise ValueError("New password must be at least 8 characters.")
    if new_password != confirm_password:
        raise ValueError("New passwords do not match.")
    if check_password_hash(account["password_hash"], new_password):
        raise ValueError("New password must be different from current password.")
    account["password_hash"] = generate_password_hash(new_password, method="scrypt")
    account["password_updated_on"] = date.today().isoformat()
    temporary = AUTH_FILE.with_suffix(".tmp")
    temporary.write_text(json.dumps(account, indent=2), encoding="utf-8")
    temporary.replace(AUTH_FILE)
    try:
        AUTH_FILE.chmod(0o600)
    except OSError:
        pass


provision_default_account()


@app.before_request
def cors_preflight():
    if request.method == "OPTIONS":
        return ("", 204)


def csrf_token() -> str:
    token = session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["csrf_token"] = token
    return token


def csrf_is_valid() -> bool:
    supplied = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token", "")
    return bool(supplied and secrets.compare_digest(supplied, session.get("csrf_token", "")))


def login_key() -> str:
    return request.remote_addr or "local"


def login_is_limited(key: str) -> bool:
    cutoff = monotonic() - LOGIN_WINDOW_SECONDS
    failures = [attempt for attempt in _login_failures.get(key, []) if attempt > cutoff]
    _login_failures[key] = failures
    return len(failures) >= LOGIN_MAX_ATTEMPTS


def record_login_failure(key: str) -> None:
    _login_failures.setdefault(key, []).append(monotonic())


def safe_redirect_target(target: str | None) -> str:
    if not target:
        return url_for("index")
    parsed = urlsplit(target)
    if parsed.netloc or parsed.scheme or not target.startswith("/") or target.startswith("//"):
        return url_for("index")
    return target


def is_authenticated() -> bool:
    return bool(session.get("username"))


def auth_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if is_authenticated():
            return view(*args, **kwargs)
        if request.path.startswith("/api/"):
            return jsonify({"error": "Authentication required."}), 401
        return redirect(url_for("login", next=request.full_path.rstrip("?")))

    return wrapped


def serialized_saved_data_write(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        with _saved_data_lock:
            return view(*args, **kwargs)

    return wrapped


def default_dates() -> tuple[str, str]:
    return "2025-06-01", "2026-05-25"


def load_latest_summary(source: str = "MT5") -> dict:
    source = normalize_source(source)
    source_path = RESULTS_DIR / source.lower() / "backtest_summary.json"
    path = source_path if source_path.exists() else RESULTS_DIR / "backtest_summary.json"
    if source != "MT5" and not source_path.exists():
        raise FileNotFoundError(f"No {source} backtest result found. Run a {source} backtest first.")
    if not path.exists():
        raise FileNotFoundError("No backtest result found. Run a backtest first.")
    return json.loads(path.read_text(encoding="utf-8"))


def clear_saved_data() -> None:
    for directory in (RESULTS_DIR, CACHE_DIR):
        if directory.exists():
            for item in directory.iterdir():
                if item.is_dir():
                    shutil.rmtree(item)
                else:
                    item.unlink()
        directory.mkdir(exist_ok=True)
    REPORT_DATA_FILE.write_text("window.BACKTEST_REPORT = null;\n", encoding="utf-8")


def clear_python_caches() -> None:
    skipped_directories = {".git", ".venv", "venv", "env", "node_modules"}
    cache_directories = {"__pycache__", ".pytest_cache", ".mypy_cache", ".ruff_cache"}
    compiled_suffixes = {".pyc", ".pyo"}
    for current_root, directories, files in os.walk(DATA_DIR, topdown=True):
        directories[:] = [name for name in directories if name not in skipped_directories]
        for name in tuple(directories):
            if name in cache_directories:
                shutil.rmtree(Path(current_root) / name)
                directories.remove(name)
        for name in files:
            if Path(name).suffix.lower() in compiled_suffixes:
                (Path(current_root) / name).unlink()


def add_sheet_rows(workbook: Workbook, title: str, rows: list[dict]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["No data"])
        return

    headers = list(rows[0].keys())
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")

    for row in rows:
        sheet.append([row.get(header) for header in headers])

    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = min(max(width, 12), 28)


def grouped_points(trades: list[dict], mode: str) -> list[dict]:
    groups: dict[str, dict] = {}
    for trade in trades:
        key = trade["trade_date"]
        if mode == "month":
            key = key[:7]
        elif mode == "year":
            key = key[:4]
        bucket = groups.setdefault(key, {"period": key, "trades": 0, "net_points": 0.0})
        bucket["trades"] += 1
        bucket["net_points"] += float(trade.get("pnl_points", 0))
    return [
        {"period": key, "trades": value["trades"], "net_points": round(value["net_points"], 2)}
        for key, value in sorted(groups.items())
    ]


def scan_values(data: dict, key: str, default: list, cast, maximum: int = 12) -> list:
    raw = data.get(key, default)
    values = raw if isinstance(raw, list) else str(raw).split(",")
    parsed = []
    for value in values:
        text = str(value).strip()
        if not text:
            continue
        item = cast(text)
        if item not in parsed:
            parsed.append(item)
    if not parsed:
        return list(default)
    if len(parsed) > maximum:
        raise ValueError(f"{key} accepts at most {maximum} values.")
    return parsed


def scan_entry_patterns(data: dict) -> list[str]:
    patterns = [value.upper() for value in scan_values(data, "entry_patterns", list(OPTIMIZER_ENTRY_PATTERNS), str, maximum=8)]
    invalid = [value for value in patterns if value not in OPTIMIZER_ENTRY_PATTERNS]
    if invalid:
        raise ValueError(f"Unsupported entry pattern: {', '.join(invalid)}")
    return patterns


def add_minutes(value, minutes: int):
    total = value.hour * 60 + value.minute + minutes
    total %= 24 * 60
    return time(total // 60, total % 60)


def minutes_of_day(value: time) -> int:
    return value.hour * 60 + value.minute


def ensure_time(value) -> time:
    if isinstance(value, time):
        return value
    return parse_time(str(value))


def minutes_after(anchor: int, value: time) -> int:
    value = ensure_time(value)
    minutes = minutes_of_day(value)
    while minutes < anchor:
        minutes += 24 * 60
    return minutes


def scan_time_profiles(data: dict) -> list[dict]:
    range_starts = scan_values(data, "range_start_values", [data.get("range_start", "08:30")], parse_time, maximum=24)
    durations = scan_values(data, "range_duration_values", [60], int, maximum=6)
    session_ends = scan_values(data, "session_end_values", [data.get("session_end", "19:30")], parse_time, maximum=24)
    setup_session_end = ensure_time(data.get("session_end", "19:30"))
    if setup_session_end not in session_ends:
        session_ends.append(setup_session_end)
    entry_cutoff_values = (
        scan_values(data, "entry_cutoff_values", [], parse_time, maximum=12)
        if str(data.get("entry_cutoff_values", "")).strip()
        else [ensure_time(data.get("entry_cutoff", "18:00"))]
    )
    profiles = []
    seen = set()
    rejected = {"range": 0, "session": 0, "entry": 0}
    first_hint = ""
    for range_start in range_starts:
        for duration in durations:
            if duration <= 0:
                raise ValueError("Range duration must be positive.")
            range_end = add_minutes(range_start, duration)
            session_start = range_end
            range_start_abs = minutes_of_day(range_start)
            range_end_abs = range_start_abs + duration
            session_start_abs = range_end_abs
            for session_end in session_ends:
                for entry_cutoff in entry_cutoff_values:
                    raw_entry_cutoff_abs = minutes_after(range_start_abs, entry_cutoff)
                    session_end_abs = minutes_after(session_start_abs + 1, session_end)
                    if raw_entry_cutoff_abs < session_start_abs:
                        entry_cutoff_abs = session_start_abs
                        effective_entry_cutoff = session_start
                    else:
                        entry_cutoff_abs = minutes_after(session_start_abs, entry_cutoff)
                        effective_entry_cutoff = entry_cutoff
                    if range_end_abs <= range_start_abs:
                        rejected["range"] += 1
                        first_hint = first_hint or f"Range {range_start.strftime('%H:%M')} + {duration} min wraps past midnight; current optimizer supports same-day range windows."
                        continue
                    if session_end_abs <= session_start_abs:
                        rejected["session"] += 1
                        first_hint = first_hint or f"Range start {range_start.strftime('%H:%M')} makes session start {session_start.strftime('%H:%M')}; Force Exit must be after this time."
                        continue
                    if entry_cutoff_abs < session_start_abs or entry_cutoff_abs >= session_end_abs:
                        rejected["entry"] += 1
                        first_hint = first_hint or f"Range start {range_start.strftime('%H:%M')} makes session start {session_start.strftime('%H:%M')}; Last Entry must be after {session_start.strftime('%H:%M')} and before {session_end.strftime('%H:%M')}."
                        continue
                    key = (
                        range_start.strftime("%H:%M"),
                        range_end.strftime("%H:%M"),
                        session_start.strftime("%H:%M"),
                        effective_entry_cutoff.strftime("%H:%M"),
                        session_end.strftime("%H:%M"),
                    )
                    if key in seen:
                        continue
                    seen.add(key)
                    profiles.append(
                        {
                            "range_start": range_start,
                            "range_end": range_end,
                            "session_start": session_start,
                            "entry_cutoff": effective_entry_cutoff,
                            "session_end": session_end,
                        }
                    )
    if not profiles:
        detail = f" Rejected windows: range={rejected['range']}, session={rejected['session']}, entry={rejected['entry']}."
        hint = f" {first_hint}" if first_hint else ""
        raise ValueError(f"No valid optimizer time windows. Check range start, cutoff and force exit values.{detail}{hint}")
    return profiles


def optimizer_payload(data: dict) -> dict:
    source = normalize_source(data.get("data_source", "DELTA"))
    symbol = str(data.get("symbol", "")).strip()
    if not symbol:
        raise ValueError("Symbol is required.")
    from_date = parse_date(data["from_date"])
    to_date = parse_date(data["to_date"])
    if from_date > to_date:
        raise ValueError("From date must be before to date.")
    available_timeframes = SOURCE_TIMEFRAMES[source]
    entry_timeframes = [value.upper() for value in scan_values(data, "entry_timeframes", available_timeframes, str)]
    trail_timeframes = [value.upper() for value in scan_values(data, "trail_timeframes", available_timeframes, str)]
    for timeframe in [*entry_timeframes, *trail_timeframes]:
        validate_source_timeframe(source, timeframe)
    entry_patterns = scan_entry_patterns(data)
    parameters = {
        "entry_buffer_pct": scan_values(data, "entry_buffer_values", [0.05, 0.20, 0.25], float),
        "stop_points": scan_values(data, "stop_points_values", [400.0, 500.0, 300.0, 200.0], float),
        "first_trail_profit": scan_values(data, "first_trail_profit_values", [400.0, 700.0, 600.0, 500.0], float),
        "first_trail_lock_loss": scan_values(data, "first_trail_lock_values", [300.0, 400.0, 200.0], float),
        "second_trail_profit": scan_values(data, "second_trail_profit_values", [700.0], float),
    }
    distance_units = {
        "stop_points_unit": normalize_distance_unit(data.get("stop_points_unit", "POINTS")),
        "first_trail_profit_unit": normalize_distance_unit(data.get("first_trail_profit_unit", "POINTS")),
        "first_trail_lock_loss_unit": normalize_distance_unit(data.get("first_trail_lock_loss_unit", "POINTS")),
        "second_trail_profit_unit": normalize_distance_unit(data.get("second_trail_profit_unit", "POINTS")),
    }
    if any(value < 0 for values in parameters.values() for value in values):
        raise ValueError("Optimizer point and percentage settings must not be negative.")
    common = {
        "symbol": symbol,
        "from_date": from_date,
        "to_date": to_date,
        "data_source": source,
    }
    time_profiles = scan_time_profiles(data)
    total = (
        len(entry_timeframes)
        * len(trail_timeframes)
        * len(parameters["entry_buffer_pct"])
        * len(parameters["stop_points"])
        * len(parameters["first_trail_profit"])
        * len(parameters["first_trail_lock_loss"])
        * len(parameters["second_trail_profit"])
        * len(entry_patterns)
        * len(time_profiles)
    )
    maximum = min(max(int(data.get("max_combinations", OPTIMIZER_MAX_COMBINATIONS)), 1), OPTIMIZER_MAX_COMBINATIONS)
    if total > maximum:
        raise ValueError(f"Scan has {total:,} combinations. Reduce values or increase limit up to {OPTIMIZER_MAX_COMBINATIONS:,}.")
    return {
        "common": common,
        "entry_timeframes": entry_timeframes,
        "trail_timeframes": trail_timeframes,
        "entry_patterns": entry_patterns,
        "time_profiles": time_profiles,
        "parameters": parameters,
        "distance_units": distance_units,
        "target_win_rate": min(max(float(data.get("target_win_rate", 70)), 0), 100),
        "minimum_trades": max(int(data.get("minimum_trades", 20)), 1),
        "result_sort": str(data.get("result_sort", "BALANCED")).upper() if str(data.get("result_sort", "BALANCED")).upper() in OPTIMIZER_RESULT_SORTS else "BALANCED",
        "stop_on_target": str(data.get("stop_on_target", "false")).lower() in {"true", "1", "on", "yes"},
        "save_full_csv": str(data.get("save_full_csv", OPTIMIZER_SAVE_FULL_CSV_DEFAULT)).lower() in {"true", "1", "on", "yes"},
        "total": total,
    }


def optimization_row(config: BacktestConfig, stats: dict, qualified: bool, tested: int, entry_pattern: str) -> dict:
    return {
        "rank": int(tested),
        "qualified": bool(qualified),
        "entry_pattern": entry_pattern,
        "timeframe": config.timeframe,
        "trail_timeframe": config.trail_timeframe,
        "range_start": config.range_start.strftime("%H:%M"),
        "range_end": config.range_end.strftime("%H:%M"),
        "session_start": config.session_start.strftime("%H:%M"),
        "entry_cutoff": config.entry_cutoff.strftime("%H:%M"),
        "session_end": config.session_end.strftime("%H:%M"),
        "entry_buffer_pct": round(float(config.entry_buffer_pct * 100), 6),
        "stop_points": float(config.stop_points),
        "stop_points_unit": config.stop_points_unit,
        "first_trail_profit": float(config.first_trail_profit),
        "first_trail_profit_unit": config.first_trail_profit_unit,
        "first_trail_lock_loss": float(config.first_trail_lock_loss),
        "first_trail_lock_loss_unit": config.first_trail_lock_loss_unit,
        "second_trail_profit": float(config.second_trail_profit),
        "second_trail_profit_unit": config.second_trail_profit_unit,
        "total_trades": int(stats["total_trades"]),
        "wins": int(stats["wins"]),
        "losses": int(stats["losses"]),
        "win_rate_pct": float(stats["win_rate_pct"]),
        "net_points": float(stats["net_points"]),
        "profit_factor": None if stats["profit_factor"] is None else float(stats["profit_factor"]),
        "max_drawdown_points": float(stats["max_drawdown_points"]),
    }


def optimizer_stats_from_metric(metric) -> dict:
    total, wins, losses, gross_profit, gross_loss, net, max_drawdown = metric
    return {
        "total_trades": int(total),
        "wins": int(wins),
        "losses": int(losses),
        "win_rate_pct": round((wins / total) * 100, 2) if total else 0.0,
        "net_points": round(float(net), 2),
        "profit_factor": round(float(gross_profit / abs(gross_loss)), 2) if gross_loss else None,
        "max_drawdown_points": round(float(max_drawdown), 2),
    }


def optimizer_metric_arrays(metrics: np.ndarray, minimum_trades: int, target_win_rate: float) -> dict[str, np.ndarray]:
    total = metrics[:, 0]
    wins = metrics[:, 1]
    win_rate = np.divide(wins * 100.0, total, out=np.zeros_like(wins), where=total > 0)
    win_rate = np.round(win_rate, 2)
    gross_profit = metrics[:, 3]
    gross_loss = metrics[:, 4]
    profit_factor = np.divide(gross_profit, np.abs(gross_loss), out=np.zeros_like(gross_profit), where=gross_loss != 0)
    profit_factor = np.round(profit_factor, 2)
    net_points = np.round(metrics[:, 5], 2)
    drawdown = np.abs(np.round(metrics[:, 6], 2))
    trades = total.astype(np.float64)
    qualified = ((trades >= minimum_trades) & (win_rate >= target_win_rate)).astype(np.float64)
    return {
        "qualified": qualified,
        "win_rate": win_rate,
        "net_points": net_points,
        "profit_factor": profit_factor,
        "trades": trades,
        "drawdown": drawdown,
    }


def optimizer_metric_sort_keys(metrics: np.ndarray, minimum_trades: int, target_win_rate: float, mode: str) -> tuple[np.ndarray, ...]:
    arrays = optimizer_metric_arrays(metrics, minimum_trades, target_win_rate)
    mode = mode if mode in OPTIMIZER_RESULT_SORTS else "BALANCED"
    if mode == "WIN_RATE":
        return (arrays["qualified"], arrays["win_rate"], arrays["net_points"], arrays["profit_factor"], arrays["trades"])
    if mode == "NET_POINTS":
        return (arrays["qualified"], arrays["net_points"], arrays["win_rate"], arrays["profit_factor"], arrays["trades"])
    if mode == "LOWEST_TRADES":
        return (arrays["qualified"], -arrays["trades"], arrays["win_rate"], arrays["net_points"], arrays["profit_factor"])
    if mode == "WIN_POINTS":
        return (arrays["qualified"], arrays["win_rate"], arrays["net_points"], -arrays["drawdown"], arrays["trades"])
    return (arrays["qualified"], arrays["win_rate"], arrays["net_points"], arrays["trades"], arrays["profit_factor"], -arrays["drawdown"])


def optimizer_top_metric_indices(metrics: np.ndarray, minimum_trades: int, target_win_rate: float, mode: str, limit: int) -> np.ndarray:
    traded_indices = np.flatnonzero(metrics[:, 0] > 0)
    if len(traded_indices) == 0:
        return np.empty(0, dtype=np.int64)
    if len(traded_indices) <= limit:
        return traded_indices
    traded_metrics = metrics[traded_indices]
    keys = optimizer_metric_sort_keys(traded_metrics, minimum_trades, target_win_rate, mode)
    order = np.lexsort(tuple(-key for key in reversed(keys)))
    return traded_indices[order[:limit]]


def optimizer_qualified_mask(metrics: np.ndarray, minimum_trades: int, target_win_rate: float) -> np.ndarray:
    arrays = optimizer_metric_arrays(metrics, minimum_trades, target_win_rate)
    return arrays["qualified"].astype(np.bool_)


def optimization_row_from_metric(
    profile: dict,
    entry_tf: str,
    trail_tf: str,
    distance_units: dict,
    parameter_values,
    metric,
    qualified: bool,
    tested: int,
    entry_pattern: str,
) -> dict:
    stats = optimizer_stats_from_metric(metric)
    return {
        "rank": int(tested),
        "qualified": bool(qualified),
        "entry_pattern": entry_pattern,
        "timeframe": entry_tf,
        "trail_timeframe": trail_tf,
        "range_start": profile["range_start"].strftime("%H:%M"),
        "range_end": profile["range_end"].strftime("%H:%M"),
        "session_start": profile["session_start"].strftime("%H:%M"),
        "entry_cutoff": profile["entry_cutoff"].strftime("%H:%M"),
        "session_end": profile["session_end"].strftime("%H:%M"),
        "entry_buffer_pct": round(float(parameter_values[0] * 100), 6),
        "stop_points": float(parameter_values[1]),
        "stop_points_unit": distance_units.get("stop_points_unit", "POINTS"),
        "first_trail_profit": float(parameter_values[2]),
        "first_trail_profit_unit": distance_units.get("first_trail_profit_unit", "POINTS"),
        "first_trail_lock_loss": float(parameter_values[3]),
        "first_trail_lock_loss_unit": distance_units.get("first_trail_lock_loss_unit", "POINTS"),
        "second_trail_profit": float(parameter_values[4]),
        "second_trail_profit_unit": distance_units.get("second_trail_profit_unit", "POINTS"),
        "total_trades": stats["total_trades"],
        "wins": stats["wins"],
        "losses": stats["losses"],
        "win_rate_pct": stats["win_rate_pct"],
        "net_points": stats["net_points"],
        "profit_factor": stats["profit_factor"],
        "max_drawdown_points": stats["max_drawdown_points"],
    }


def optimizer_parameter_chunks(parameters: dict, chunk_size: int):
    chunk = np.empty((chunk_size, 5), dtype=np.float64)
    index = 0
    for buffer_pct, stop, first_profit, first_lock, second_profit in product(
        parameters["entry_buffer_pct"],
        parameters["stop_points"],
        parameters["first_trail_profit"],
        parameters["first_trail_lock_loss"],
        parameters["second_trail_profit"],
    ):
        chunk[index, 0] = buffer_pct / 100
        chunk[index, 1] = stop
        chunk[index, 2] = first_profit
        chunk[index, 3] = first_lock
        chunk[index, 4] = second_profit
        index += 1
        if index == chunk_size:
            yield chunk
            chunk = np.empty((chunk_size, 5), dtype=np.float64)
            index = 0
    if index:
        yield chunk[:index].copy()


def optimizer_rank_key(row: dict, mode: str = "BALANCED") -> tuple:
    mode = mode if mode in OPTIMIZER_RESULT_SORTS else "BALANCED"
    qualified = bool(row.get("qualified"))
    win_rate = float(row.get("win_rate_pct") or 0)
    net_points = float(row.get("net_points") or 0)
    trades = int(row.get("total_trades") or 0)
    profit_factor = float(row.get("profit_factor") or 0)
    drawdown = abs(float(row.get("max_drawdown_points") or 0))
    if mode == "WIN_RATE":
        return (qualified, win_rate, net_points, profit_factor, trades)
    if mode == "NET_POINTS":
        return (qualified, net_points, win_rate, profit_factor, trades)
    if mode == "LOWEST_TRADES":
        return (qualified, -trades, win_rate, net_points, profit_factor)
    if mode == "WIN_POINTS":
        return (qualified, win_rate, net_points, -drawdown, trades)
    return (qualified, win_rate, net_points, trades, profit_factor, -drawdown)


def safe_optimizer_part(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value).strip().upper()) or "UNKNOWN"


def optimizer_scan_dir(source: str, symbol: str, scan_id: str) -> Path:
    return OPTIMIZER_RUNS_DIR / normalize_source(source).lower() / safe_optimizer_part(symbol) / safe_optimizer_part(scan_id)


def optimizer_output_path(source: str, symbol: str, scan_id: str) -> Path:
    return optimizer_scan_dir(source, symbol, scan_id) / "scan.json"


def optimizer_csv_path(source: str, symbol: str, scan_id: str) -> Path:
    return optimizer_scan_dir(source, symbol, scan_id) / "results.csv"


def optimizer_db_connection() -> sqlite3.Connection:
    OPTIMIZER_DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(OPTIMIZER_DB_FILE, timeout=30.0)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute("PRAGMA synchronous=NORMAL")
    connection.execute("PRAGMA temp_store=MEMORY")
    connection.execute("PRAGMA foreign_keys=ON")
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS optimizer_scans (
            source TEXT NOT NULL,
            symbol_key TEXT NOT NULL,
            symbol TEXT NOT NULL,
            scan_id TEXT NOT NULL,
            engine_version INTEGER NOT NULL,
            status TEXT NOT NULL,
            message TEXT,
            created_at TEXT,
            updated_at TEXT,
            tested INTEGER NOT NULL DEFAULT 0,
            total INTEGER NOT NULL DEFAULT 0,
            progress_pct REAL NOT NULL DEFAULT 0,
            target_win_rate REAL NOT NULL DEFAULT 0,
            minimum_trades INTEGER NOT NULL DEFAULT 1,
            result_sort TEXT NOT NULL DEFAULT 'BALANCED',
            save_full_csv INTEGER NOT NULL DEFAULT 0,
            result_count INTEGER NOT NULL DEFAULT 0,
            results_truncated INTEGER NOT NULL DEFAULT 0,
            target_found INTEGER NOT NULL DEFAULT 0,
            stopped_early INTEGER NOT NULL DEFAULT 0,
            best_win_rate REAL,
            elapsed_seconds REAL,
            scan_config TEXT NOT NULL,
            PRIMARY KEY (source, symbol_key, scan_id)
        );
        CREATE TABLE IF NOT EXISTS optimizer_results (
            source TEXT NOT NULL,
            symbol_key TEXT NOT NULL,
            scan_id TEXT NOT NULL,
            rank INTEGER NOT NULL,
            qualified INTEGER NOT NULL DEFAULT 0,
            entry_pattern TEXT,
            timeframe TEXT,
            trail_timeframe TEXT,
            range_start TEXT,
            range_end TEXT,
            session_start TEXT,
            entry_cutoff TEXT,
            session_end TEXT,
            entry_buffer_pct REAL,
            stop_points REAL,
            stop_points_unit TEXT,
            first_trail_profit REAL,
            first_trail_profit_unit TEXT,
            first_trail_lock_loss REAL,
            first_trail_lock_loss_unit TEXT,
            second_trail_profit REAL,
            second_trail_profit_unit TEXT,
            total_trades INTEGER,
            wins INTEGER,
            losses INTEGER,
            win_rate_pct REAL,
            net_points REAL,
            profit_factor REAL,
            max_drawdown_points REAL,
            PRIMARY KEY (source, symbol_key, scan_id, rank),
            FOREIGN KEY (source, symbol_key, scan_id)
                REFERENCES optimizer_scans(source, symbol_key, scan_id)
                ON DELETE CASCADE
        );
        CREATE INDEX IF NOT EXISTS idx_optimizer_scans_lookup
            ON optimizer_scans(source, symbol_key, created_at DESC);
        """
    )
    return connection


def optimizer_symbol_key(symbol: str) -> str:
    return safe_optimizer_part(symbol)


def db_int(value, default: int = 0) -> int:
    if value is None or value == "":
        return default
    return int(value)


def db_float(value, default: float | None = 0.0) -> float | None:
    if value is None or value == "":
        return default
    return float(value)


def db_bool(value) -> int:
    return 1 if bool(value) else 0


def scan_from_db_row(row: sqlite3.Row, results: list[dict] | None = None) -> dict:
    scan_config = json.loads(row["scan_config"])
    record = {
        "scan_id": row["scan_id"],
        "source": row["source"],
        "symbol": row["symbol"],
        "status": row["status"],
        "message": row["message"] or "",
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
        "tested": int(row["tested"] or 0),
        "total": int(row["total"] or 0),
        "progress_pct": float(row["progress_pct"] or 0),
        "target_win_rate": float(row["target_win_rate"] or 0),
        "minimum_trades": int(row["minimum_trades"] or 1),
        "result_sort": row["result_sort"] or "BALANCED",
        "save_full_csv": bool(row["save_full_csv"]),
        "result_count": int(row["result_count"] or 0),
        "results_truncated": bool(row["results_truncated"]),
        "target_found": bool(row["target_found"]),
        "stopped_early": bool(row["stopped_early"]),
        "scan_config": scan_config,
        "results": results or [],
    }
    if row["best_win_rate"] is not None:
        record["best_win_rate"] = float(row["best_win_rate"])
    if row["elapsed_seconds"] is not None:
        record["elapsed_seconds"] = float(row["elapsed_seconds"])
    return record


def optimizer_result_from_db_row(row: sqlite3.Row) -> dict:
    return {
        "rank": int(row["rank"]),
        "qualified": bool(row["qualified"]),
        "entry_pattern": row["entry_pattern"],
        "timeframe": row["timeframe"],
        "trail_timeframe": row["trail_timeframe"],
        "range_start": row["range_start"],
        "range_end": row["range_end"],
        "session_start": row["session_start"],
        "entry_cutoff": row["entry_cutoff"],
        "session_end": row["session_end"],
        "entry_buffer_pct": db_float(row["entry_buffer_pct"]),
        "stop_points": db_float(row["stop_points"]),
        "stop_points_unit": row["stop_points_unit"],
        "first_trail_profit": db_float(row["first_trail_profit"]),
        "first_trail_profit_unit": row["first_trail_profit_unit"],
        "first_trail_lock_loss": db_float(row["first_trail_lock_loss"]),
        "first_trail_lock_loss_unit": row["first_trail_lock_loss_unit"],
        "second_trail_profit": db_float(row["second_trail_profit"]),
        "second_trail_profit_unit": row["second_trail_profit_unit"],
        "total_trades": db_int(row["total_trades"]),
        "wins": db_int(row["wins"]),
        "losses": db_int(row["losses"]),
        "win_rate_pct": db_float(row["win_rate_pct"]),
        "net_points": db_float(row["net_points"]),
        "profit_factor": db_float(row["profit_factor"], None),
        "max_drawdown_points": db_float(row["max_drawdown_points"]),
    }


def save_optimizer_scan_db(scan: dict) -> None:
    config = scan.get("scan_config", {})
    source = normalize_source(scan.get("source") or config.get("data_source", "DELTA"))
    symbol = str(scan.get("symbol") or config.get("symbol", "")).strip()
    symbol_key = optimizer_symbol_key(symbol)
    results = scan.get("results") or []
    with _optimizer_db_lock:
        connection = optimizer_db_connection()
        try:
            with connection:
                connection.execute(
                    """
                    INSERT INTO optimizer_scans (
                        source, symbol_key, symbol, scan_id, engine_version, status, message,
                        created_at, updated_at, tested, total, progress_pct, target_win_rate,
                        minimum_trades, result_sort, save_full_csv, result_count,
                        results_truncated, target_found, stopped_early, best_win_rate,
                        elapsed_seconds, scan_config
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(source, symbol_key, scan_id) DO UPDATE SET
                        symbol=excluded.symbol,
                        engine_version=excluded.engine_version,
                        status=excluded.status,
                        message=excluded.message,
                        created_at=excluded.created_at,
                        updated_at=excluded.updated_at,
                        tested=excluded.tested,
                        total=excluded.total,
                        progress_pct=excluded.progress_pct,
                        target_win_rate=excluded.target_win_rate,
                        minimum_trades=excluded.minimum_trades,
                        result_sort=excluded.result_sort,
                        save_full_csv=excluded.save_full_csv,
                        result_count=excluded.result_count,
                        results_truncated=excluded.results_truncated,
                        target_found=excluded.target_found,
                        stopped_early=excluded.stopped_early,
                        best_win_rate=excluded.best_win_rate,
                        elapsed_seconds=excluded.elapsed_seconds,
                        scan_config=excluded.scan_config
                    """,
                    (
                        source,
                        symbol_key,
                        symbol,
                        scan["scan_id"],
                        int(config.get("engine_version", OPTIMIZER_ENGINE_VERSION)),
                        scan.get("status", "queued"),
                        scan.get("message", ""),
                        scan.get("created_at"),
                        scan.get("updated_at"),
                        db_int(scan.get("tested")),
                        db_int(scan.get("total")),
                        db_float(scan.get("progress_pct")),
                        db_float(scan.get("target_win_rate")),
                        db_int(scan.get("minimum_trades"), 1),
                        scan.get("result_sort", "BALANCED"),
                        db_bool(scan.get("save_full_csv")),
                        db_int(scan.get("result_count")),
                        db_bool(scan.get("results_truncated")),
                        db_bool(scan.get("target_found")),
                        db_bool(scan.get("stopped_early")),
                        db_float(scan.get("best_win_rate"), None),
                        db_float(scan.get("elapsed_seconds"), None),
                        json.dumps(config, separators=(",", ":")),
                    ),
                )
                connection.execute(
                    "DELETE FROM optimizer_results WHERE source = ? AND symbol_key = ? AND scan_id = ?",
                    (source, symbol_key, scan["scan_id"]),
                )
                connection.executemany(
                    """
                    INSERT INTO optimizer_results (
                        source, symbol_key, scan_id, rank, qualified, entry_pattern, timeframe,
                        trail_timeframe, range_start, range_end, session_start, entry_cutoff,
                        session_end, entry_buffer_pct, stop_points, stop_points_unit,
                        first_trail_profit, first_trail_profit_unit, first_trail_lock_loss,
                        first_trail_lock_loss_unit, second_trail_profit, second_trail_profit_unit,
                        total_trades, wins, losses, win_rate_pct, net_points, profit_factor,
                        max_drawdown_points
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    [
                        (
                            source,
                            symbol_key,
                            scan["scan_id"],
                            db_int(row.get("rank"), index),
                            db_bool(row.get("qualified")),
                            row.get("entry_pattern"),
                            row.get("timeframe"),
                            row.get("trail_timeframe"),
                            row.get("range_start"),
                            row.get("range_end"),
                            row.get("session_start"),
                            row.get("entry_cutoff"),
                            row.get("session_end"),
                            db_float(row.get("entry_buffer_pct")),
                            db_float(row.get("stop_points")),
                            row.get("stop_points_unit"),
                            db_float(row.get("first_trail_profit")),
                            row.get("first_trail_profit_unit"),
                            db_float(row.get("first_trail_lock_loss")),
                            row.get("first_trail_lock_loss_unit"),
                            db_float(row.get("second_trail_profit")),
                            row.get("second_trail_profit_unit"),
                            db_int(row.get("total_trades")),
                            db_int(row.get("wins")),
                            db_int(row.get("losses")),
                            db_float(row.get("win_rate_pct")),
                            db_float(row.get("net_points")),
                            db_float(row.get("profit_factor"), None),
                            db_float(row.get("max_drawdown_points")),
                        )
                        for index, row in enumerate(results, start=1)
                    ],
                )
        finally:
            connection.close()


def optimizer_results_from_db(connection: sqlite3.Connection, source: str, symbol_key: str, scan_id: str) -> list[dict]:
    rows = connection.execute(
        """
        SELECT * FROM optimizer_results
        WHERE source = ? AND symbol_key = ? AND scan_id = ?
        ORDER BY rank ASC
        """,
        (source, symbol_key, scan_id),
    ).fetchall()
    return [optimizer_result_from_db_row(row) for row in rows]


def load_optimizer_scan_db(source: str, symbol: str, scan_id: str) -> dict | None:
    source = normalize_source(source)
    symbol_key = optimizer_symbol_key(symbol)
    connection = optimizer_db_connection()
    try:
        row = connection.execute(
            """
            SELECT * FROM optimizer_scans
            WHERE source = ? AND symbol_key = ? AND scan_id = ? AND engine_version = ?
            """,
            (source, symbol_key, scan_id, OPTIMIZER_ENGINE_VERSION),
        ).fetchone()
        if row is None:
            return None
        results = optimizer_results_from_db(connection, source, symbol_key, scan_id)
        return scan_from_db_row(row, results)
    finally:
        connection.close()


def saved_optimizer_scans_db(source: str, symbol: str, include_results: bool = True) -> list[dict]:
    source = normalize_source(source)
    symbol_key = optimizer_symbol_key(symbol)
    connection = optimizer_db_connection()
    try:
        rows = connection.execute(
            """
            SELECT * FROM optimizer_scans
            WHERE source = ? AND symbol_key = ? AND engine_version = ?
            ORDER BY created_at DESC
            LIMIT 50
            """,
            (source, symbol_key, OPTIMIZER_ENGINE_VERSION),
        ).fetchall()
        records = []
        for row in rows:
            results = optimizer_results_from_db(connection, source, symbol_key, row["scan_id"]) if include_results else []
            records.append(scan_from_db_row(row, results))
        return records
    finally:
        connection.close()


def legacy_saved_optimizer_scans(source: str, symbol: str, include_results: bool = True) -> list[dict]:
    directory = OPTIMIZER_RUNS_DIR / normalize_source(source).lower() / safe_optimizer_part(symbol)
    if not directory.exists():
        return []
    records = []
    for path in directory.glob("*/scan.json"):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if int(data.get("scan_config", {}).get("engine_version", 0)) != OPTIMIZER_ENGINE_VERSION:
                continue
            if not include_results:
                data["results"] = []
            records.append(data)
        except (OSError, json.JSONDecodeError):
            continue
    return records


def load_legacy_optimizer_scan(source: str, symbol: str, scan_id: str) -> dict | None:
    path = optimizer_output_path(source, symbol, scan_id)
    if not path.exists():
        return None
    result = json.loads(path.read_text(encoding="utf-8"))
    if int(result.get("scan_config", {}).get("engine_version", 0)) != OPTIMIZER_ENGINE_VERSION:
        raise ValueError("This optimizer scan was created by an older engine. Run a fresh scan.")
    return result


def load_saved_optimizer_scan(source: str, symbol: str, scan_id: str) -> dict | None:
    return load_optimizer_scan_db(source, symbol, scan_id) or load_legacy_optimizer_scan(source, symbol, scan_id)


def optimizer_scan_config(payload: dict) -> dict:
    common = payload["common"]
    return {
        "engine_version": OPTIMIZER_ENGINE_VERSION,
        "symbol": common["symbol"],
        "data_source": common["data_source"],
        "from_date": common["from_date"].isoformat(),
        "to_date": common["to_date"].isoformat(),
        "time_profiles": [
            {
                "range_start": profile["range_start"].strftime("%H:%M"),
                "range_end": profile["range_end"].strftime("%H:%M"),
                "session_start": profile["session_start"].strftime("%H:%M"),
                "entry_cutoff": profile["entry_cutoff"].strftime("%H:%M"),
                "session_end": profile["session_end"].strftime("%H:%M"),
            }
            for profile in payload["time_profiles"]
        ],
        "entry_timeframes": payload["entry_timeframes"],
        "trail_timeframes": payload["trail_timeframes"],
        "entry_patterns": payload["entry_patterns"],
        "parameters": payload["parameters"],
        "distance_units": payload.get("distance_units", {}),
        "target_win_rate": payload["target_win_rate"],
        "minimum_trades": payload["minimum_trades"],
        "result_sort": payload.get("result_sort", "BALANCED"),
        "stop_on_target": payload["stop_on_target"],
        "save_full_csv": payload.get("save_full_csv", False),
    }


def optimizer_public_result(job: dict, rows: list[dict]) -> dict:
    sort_mode = job.get("result_sort", "BALANCED")
    ranked = sorted(
        rows,
        key=lambda row: optimizer_rank_key(row, sort_mode),
        reverse=True,
    )
    for rank, row in enumerate(ranked, start=1):
        row["rank"] = rank
    result = dict(job)
    result["result_count"] = len(ranked)
    result["results_truncated"] = len(ranked) > OPTIMIZER_VISIBLE_RESULTS
    result["results"] = ranked[:OPTIMIZER_VISIBLE_RESULTS]
    return result


def save_optimizer_output(
    job: dict,
    payload: dict,
    rows: list[dict],
    csv_rows: list[dict] | None = None,
    replace_csv: bool = False,
    rows_are_ranked: bool = False,
    total_result_count: int | None = None,
) -> dict:
    common = payload["common"]
    if rows_are_ranked:
        stored = dict(job)
        stored["result_count"] = len(rows)
        stored["results_truncated"] = len(rows) > OPTIMIZER_VISIBLE_RESULTS
        stored["results"] = rows[:OPTIMIZER_VISIBLE_RESULTS]
    else:
        stored = optimizer_public_result(job, rows)
    if total_result_count is not None:
        stored["result_count"] = total_result_count
        stored["results_truncated"] = total_result_count > OPTIMIZER_VISIBLE_RESULTS
    stored["scan_config"] = optimizer_scan_config(payload)
    save_optimizer_scan_db(stored)
    csv_path = optimizer_csv_path(common["data_source"], common["symbol"], job["scan_id"])
    if csv_rows:
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        if replace_csv:
            csv_temp = csv_path.with_suffix(".tmp")
            handle = csv_temp.open("w", newline="", encoding="utf-8")
        else:
            csv_temp = None
            handle = csv_path.open("a", newline="", encoding="utf-8")
        with handle:
            writer = csv.DictWriter(handle, fieldnames=list(csv_rows[0].keys()))
            if replace_csv or not csv_path.exists() or csv_path.stat().st_size == 0:
                writer.writeheader()
            writer.writerows(csv_rows)
        if csv_temp:
            csv_temp.replace(csv_path)
    return stored


def saved_optimizer_scans(source: str, symbol: str, include_results: bool = True) -> list[dict]:
    records_by_scan_id = {row["scan_id"]: row for row in saved_optimizer_scans_db(source, symbol, include_results)}
    for row in legacy_saved_optimizer_scans(source, symbol, include_results):
        records_by_scan_id.setdefault(row["scan_id"], row)
    records = list(records_by_scan_id.values())
    records.sort(key=lambda item: item.get("created_at", ""), reverse=True)
    return records


def public_saved_optimizer_scan(row: dict) -> dict:
    active = row.get("scan_id") in _optimizer_jobs
    status = row.get("status")
    if status in {"queued", "running"} and not active:
        status = "interrupted"
    return {
        "scan_id": row["scan_id"],
        "status": status,
        "message": "Scan was interrupted. Saved partial results are available." if status == "interrupted" else row.get("message", ""),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "tested": row.get("tested", 0),
        "total": row.get("total", 0),
        "target_found": row.get("target_found", False),
        "best_win_rate": row.get("best_win_rate"),
        "scan_config": row.get("scan_config", {}),
    }


def execute_optimizer(job_id: str, payload: dict) -> None:
    started = monotonic()
    tested = 0
    results = []
    csv_buffer = []
    csv_started = False
    target_found = False
    frames: dict[tuple[str, bool], object] = {}
    layouts = {}
    contexts = {}
    common = payload["common"]
    distance_units = payload.get("distance_units", {})
    debug_enabled = str(payload.get("debug", "")).lower() in {"1", "true", "yes", "on"}
    m5_derived_timeframes = {"M5", "M10", "M15", "M30", "H1"}
    use_m5_derived_frames = (
        common["data_source"] == "MT5"
        and set(payload["entry_timeframes"]).issubset(m5_derived_timeframes)
        and set(payload["trail_timeframes"]).issubset(m5_derived_timeframes)
    )
    has_overnight_windows = any(
        ensure_time(profile["session_end"]) <= ensure_time(profile["session_start"])
        or ensure_time(profile["entry_cutoff"]) < ensure_time(profile["session_start"])
        for profile in payload["time_profiles"]
    )

    def checkpoint(csv_rows: list[dict] | None = None, replace_csv: bool = False, **changes) -> None:
        def persist() -> None:
            changes["updated_at"] = datetime.now(timezone.utc).isoformat()
            with _optimizer_lock:
                _optimizer_jobs[job_id].update(changes)
                current_job = dict(_optimizer_jobs[job_id])
            stored = save_optimizer_output(
                current_job,
                payload,
                results,
                csv_rows=csv_rows,
                replace_csv=replace_csv,
                total_result_count=tested,
            )
            with _optimizer_lock:
                _optimizer_jobs[job_id].update(
                    {
                        "result_count": stored["result_count"],
                        "results_truncated": stored["results_truncated"],
                        "results": stored["results"],
                    }
                )

        if changes.get("status") in {"completed", "error"}:
            with _saved_data_lock:
                persist()
        else:
            persist()

    def update_progress(**changes) -> None:
        changes["updated_at"] = datetime.now(timezone.utc).isoformat()
        changes.setdefault("elapsed_seconds", round(monotonic() - started, 2))
        changes.setdefault("tested", tested)
        changes.setdefault("progress_pct", round((tested / payload["total"]) * 100, 1))
        if results:
            sort_mode = payload.get("result_sort", "BALANCED")
            key = lambda row: optimizer_rank_key(row, sort_mode)
            ranked = heapq.nlargest(
                OPTIMIZER_VISIBLE_RESULTS,
                results,
                key=key,
            )
            for rank, row in enumerate(ranked[:OPTIMIZER_VISIBLE_RESULTS], start=1):
                row["rank"] = rank
            changes["result_count"] = tested
            changes["results_truncated"] = tested > OPTIMIZER_VISIBLE_RESULTS
            changes["results"] = ranked
        with _optimizer_lock:
            _optimizer_jobs[job_id].update(changes)

    def debug_progress(message: str) -> None:
        if debug_enabled:
            update_progress(status="running", message=message)

    def resample_m5_frame(frame, timeframe: str):
        timeframe = timeframe.upper()
        if timeframe == "M5":
            return frame.copy()
        rule = f"{TIMEFRAME_MINUTES[timeframe]}min"
        indexed = frame.sort_values("time_ist").set_index("time_ist")
        resampled = indexed.resample(rule, origin="start_day", label="left", closed="left").agg(
            open=("open", "first"),
            high=("high", "max"),
            low=("low", "min"),
            close=("close", "last"),
            tick_volume=("tick_volume", "sum"),
        )
        resampled = resampled.dropna(subset=["open", "high", "low", "close"]).reset_index()
        resampled["trade_date"] = resampled["time_ist"].dt.date
        return resampled[["time_ist", "trade_date", "open", "high", "low", "close", "tick_volume"]].copy()

    def base_m5_frame(padded: bool):
        key = ("M5_BASE", True)
        if key not in frames:
            frame_config = BacktestConfig(
                symbol=common["symbol"],
                from_date=common["from_date"] - timedelta(days=1),
                to_date=common["to_date"] + timedelta(days=1) if has_overnight_windows else common["to_date"],
                data_source=common["data_source"],
                timeframe="M5",
            )
            debug_progress(f"DEBUG fetch MT5 M5 candles: {frame_config.from_date} to {frame_config.to_date}")
            frames[key] = fetch_source_rates(frame_config)
            debug_progress(f"DEBUG fetched M5 candles: {len(frames[key])} rows")
        if padded:
            return frames[key]
        return frames[key][frames[key]["trade_date"] >= common["from_date"]].copy()

    def candle_frame(timeframe: str, padded: bool):
        key = (timeframe, padded)
        if key not in frames:
            if use_m5_derived_frames:
                debug_progress(f"DEBUG build {timeframe}{' padded' if padded else ''} candles from M5 cache")
                frames[key] = resample_m5_frame(base_m5_frame(padded), timeframe)
                debug_progress(f"DEBUG ready {timeframe}{' padded' if padded else ''}: {len(frames[key])} rows")
            else:
                frame_config = BacktestConfig(
                    symbol=common["symbol"],
                    from_date=common["from_date"] - timedelta(days=1) if padded else common["from_date"],
                    to_date=common["to_date"] + timedelta(days=1) if has_overnight_windows else common["to_date"],
                    data_source=common["data_source"],
                    timeframe=timeframe,
                )
                debug_progress(f"DEBUG fetch {timeframe}{' padded' if padded else ''} candles from source")
                frames[key] = fetch_source_rates(frame_config)
                debug_progress(f"DEBUG ready {timeframe}{' padded' if padded else ''}: {len(frames[key])} rows")
        return frames[key]

    def execution_frame(padded: bool):
        key = ("M1_EXECUTION", padded)
        if key not in frames:
            frame_config = BacktestConfig(
                symbol=common["symbol"],
                from_date=common["from_date"] - timedelta(days=1) if padded else common["from_date"],
                to_date=common["to_date"] + timedelta(days=1) if has_overnight_windows else common["to_date"],
                data_source=common["data_source"],
                timeframe="M1",
            )
            debug_progress(f"DEBUG fetch M1 execution candles: {frame_config.from_date} to {frame_config.to_date}")
            frames[key] = fetch_source_rates(frame_config)
            debug_progress(f"DEBUG ready M1 execution candles: {len(frames[key])} rows")
        return frames[key]

    def time_profile_key(profile: dict) -> tuple[str, str, str, str, str]:
        return (
            profile["range_start"].strftime("%H:%M"),
            profile["range_end"].strftime("%H:%M"),
            profile["session_start"].strftime("%H:%M"),
            profile["entry_cutoff"].strftime("%H:%M"),
            profile["session_end"].strftime("%H:%M"),
        )

    def scan_context(entry_timeframe: str, trail_timeframe: str, profile: dict):
        profile["range_start"] = ensure_time(profile["range_start"])
        profile["range_end"] = ensure_time(profile["range_end"])
        profile["session_start"] = ensure_time(profile["session_start"])
        profile["entry_cutoff"] = ensure_time(profile["entry_cutoff"])
        profile["session_end"] = ensure_time(profile["session_end"])
        base_profile_key = time_profile_key(profile)
        pair_key = (entry_timeframe, trail_timeframe, base_profile_key)
        if pair_key not in contexts:
            entry_df = candle_frame(entry_timeframe, False)
            layout_key = (entry_timeframe, base_profile_key)
            if layout_key not in layouts:
                layout_config = BacktestConfig(**common, **profile, timeframe=entry_timeframe, trail_timeframe=trail_timeframe)
                if len(layouts) % 25 == 0:
                    update_progress(
                        status="running",
                        message=f"Preparing layouts... {len(layouts)} built, {len(contexts)} contexts ready",
                    )
                if entry_timeframe == "M1":
                    layouts[layout_key] = build_entry_layout(entry_df, layout_config)
                else:
                    layouts[layout_key] = build_execution_layout(entry_df, execution_frame(False), layout_config)
            trail_df = entry_df if trail_timeframe == entry_timeframe else candle_frame(trail_timeframe, True)
            if len(contexts) % 50 == 0:
                update_progress(
                    status="running",
                    message=f"Preparing trail contexts... {len(contexts)} ready",
                )
            contexts[pair_key] = build_scan_context(layouts[layout_key], trail_df, trail_timeframe)
        return contexts[pair_key]

    try:
        update_progress(
            status="running",
            message="Loading M5 candles once and building selected scan timeframes..." if use_m5_derived_frames else "Loading candle data and scanning settings...",
        )
        stop_scan = False
        next_progress = OPTIMIZER_PROGRESS_EVERY
        next_checkpoint = OPTIMIZER_CHECKPOINT_EVERY
        last_progress = monotonic()
        for profile in payload["time_profiles"]:
            if stop_scan:
                break
            for entry_pattern in payload["entry_patterns"]:
                side_filter = OPTIMIZER_ENTRY_PATTERNS[entry_pattern]
                if stop_scan:
                    break
                for entry_tf in payload["entry_timeframes"]:
                    if stop_scan:
                        break
                    for trail_tf in payload["trail_timeframes"]:
                        if tested == 0:
                            update_progress(
                                tested=tested,
                                progress_pct=0.0,
                                elapsed_seconds=round(monotonic() - started, 2),
                                message=f"Preparing scan contexts for {entry_tf}/{trail_tf}...",
                            )
                        context = scan_context(entry_tf, trail_tf, profile)
                        for parameter_chunk in optimizer_parameter_chunks(payload["parameters"], OPTIMIZER_CHECKPOINT_BATCH):
                            metrics = evaluate_scan_batch_metrics(
                                context,
                                parameter_chunk,
                                side_filter=side_filter,
                                stop_points_unit=distance_units.get("stop_points_unit", "POINTS"),
                                first_trail_profit_unit=distance_units.get("first_trail_profit_unit", "POINTS"),
                                first_trail_lock_loss_unit=distance_units.get("first_trail_lock_loss_unit", "POINTS"),
                                second_trail_profit_unit=distance_units.get("second_trail_profit_unit", "POINTS"),
                            )
                            qualified_mask = optimizer_qualified_mask(metrics, payload["minimum_trades"], payload["target_win_rate"])
                            processed_count = len(metrics)
                            if payload["stop_on_target"]:
                                target_indices = np.flatnonzero(qualified_mask)
                                if len(target_indices):
                                    processed_count = int(target_indices[0]) + 1
                                    stop_scan = True
                            if processed_count != len(metrics):
                                metrics = metrics[:processed_count]
                                parameter_chunk = parameter_chunk[:processed_count]
                                qualified_mask = qualified_mask[:processed_count]
                            tested_start = tested
                            tested += processed_count
                            chunk_rows = []
                            if payload.get("save_full_csv"):
                                row_indices = range(processed_count)
                            else:
                                row_indices = optimizer_top_metric_indices(
                                    metrics,
                                    payload["minimum_trades"],
                                    payload["target_win_rate"],
                                    payload.get("result_sort", "BALANCED"),
                                    OPTIMIZER_VISIBLE_RESULTS,
                                )
                            for local_index in row_indices:
                                row = optimization_row_from_metric(
                                    profile,
                                    entry_tf,
                                    trail_tf,
                                    distance_units,
                                    parameter_chunk[int(local_index)],
                                    metrics[int(local_index)],
                                    bool(qualified_mask[int(local_index)]),
                                    tested_start + int(local_index) + 1,
                                    entry_pattern,
                                )
                                chunk_rows.append(row)
                                if payload.get("save_full_csv"):
                                    csv_buffer.append(row)
                            if chunk_rows:
                                results[:] = heapq.nlargest(
                                    OPTIMIZER_VISIBLE_RESULTS,
                                    [*results, *chunk_rows],
                                    key=lambda row: optimizer_rank_key(row, payload.get("result_sort", "BALANCED")),
                                )
                            target_found = target_found or bool(np.any(qualified_mask))
                            now = monotonic()
                            report_progress = (
                                stop_scan
                                or tested >= payload["total"]
                                or tested >= next_progress
                                or now - last_progress >= 1.0
                            )
                            if report_progress:
                                update_progress(
                                    tested=tested,
                                    progress_pct=round((tested / payload["total"]) * 100, 1),
                                    best_win_rate=max((row["win_rate_pct"] for row in results), default=0.0),
                                    elapsed_seconds=round(monotonic() - started, 2),
                                    message=f"Tested {tested} of {payload['total']} settings",
                                )
                                last_progress = now
                                while next_progress <= tested:
                                    next_progress += OPTIMIZER_PROGRESS_EVERY
                            if stop_scan or tested >= payload["total"] or tested >= next_checkpoint:
                                checkpoint(
                                    csv_rows=csv_buffer if payload.get("save_full_csv") else None,
                                    replace_csv=not csv_started,
                                    status="running",
                                    tested=tested,
                                    progress_pct=round((tested / payload["total"]) * 100, 1),
                                    best_win_rate=max((row["win_rate_pct"] for row in results), default=0.0),
                                    elapsed_seconds=round(monotonic() - started, 2),
                                    message=f"Tested {tested} of {payload['total']} settings",
                                )
                                csv_buffer.clear()
                                if payload.get("save_full_csv"):
                                    csv_started = True
                                while next_checkpoint <= tested:
                                    next_checkpoint += OPTIMIZER_CHECKPOINT_EVERY
                            if stop_scan:
                                break
                        if stop_scan:
                            break
            layouts.clear()
            contexts.clear()
        found = target_found
        final = {
            "status": "completed",
            "message": "Target reached." if found else "Scan completed without reaching target.",
            "tested": tested,
            "total": payload["total"],
            "progress_pct": 100.0 if found and payload["stop_on_target"] else round((tested / payload["total"]) * 100, 1),
            "stopped_early": found and payload["stop_on_target"] and tested < payload["total"],
            "target_found": found,
            "target_win_rate": payload["target_win_rate"],
            "minimum_trades": payload["minimum_trades"],
            "elapsed_seconds": round(monotonic() - started, 2),
        }
        ranked_rows = sorted(
            results,
            key=lambda row: optimizer_rank_key(row, payload.get("result_sort", "BALANCED")),
            reverse=True,
        )
        for rank, row in enumerate(ranked_rows, start=1):
            row["rank"] = rank
        results[:] = ranked_rows
        final_csv_rows = csv_buffer if payload.get("save_full_csv") else ranked_rows
        checkpoint(csv_rows=final_csv_rows, replace_csv=not csv_started or not payload.get("save_full_csv"), **final)
    except Exception as exc:
        checkpoint(
            csv_rows=csv_buffer,
            replace_csv=not csv_started,
            status="error",
            message=str(exc),
            tested=tested,
            progress_pct=round((tested / payload["total"]) * 100, 1),
            elapsed_seconds=round(monotonic() - started, 2),
        )


def mark_optimizer_submit_error(job_id: str, future) -> None:
    try:
        future.result()
    except Exception as exc:
        with _optimizer_lock:
            job = _optimizer_jobs.get(job_id)
            if job and job.get("status") in {"queued", "running"}:
                job.update(
                    {
                        "status": "error",
                        "message": str(exc),
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }
                )


@app.after_request
def add_headers(response):
    response.headers["Cache-Control"] = "no-store"
    response.headers["Referrer-Policy"] = "same-origin"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    origin = request.headers.get("Origin")
    if origin in {"null", "http://127.0.0.1:5000", "http://localhost:5000"}:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-CSRF-Token"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response


@app.route("/login", methods=["GET", "POST"])
def login():
    if is_authenticated():
        return redirect(url_for("index"))
    error = None
    target = request.form.get("next") or request.args.get("next", "")
    if request.method == "POST":
        key = login_key()
        if not csrf_is_valid():
            error = "This form expired. Please try again."
        elif login_is_limited(key):
            error = "Too many failed attempts. Wait five minutes before trying again."
        else:
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            try:
                account = load_account()
            except (OSError, ValueError, json.JSONDecodeError):
                return render_template("login.html", csrf_token=csrf_token(), error="Account configuration could not be read.", next=target), 500
            valid_username = secrets.compare_digest(username, account["username"])
            valid_password = check_password_hash(account["password_hash"], password)
            if valid_username and valid_password:
                _login_failures.pop(key, None)
                session.clear()
                session.permanent = True
                session["username"] = account["username"]
                csrf_token()
                return redirect(safe_redirect_target(target))
            record_login_failure(key)
            error = "Invalid user ID or password."
    return render_template("login.html", csrf_token=csrf_token(), error=error, next=target)


@app.post("/logout")
@auth_required
def logout():
    if not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    session.clear()
    return redirect(url_for("login"))


@app.get("/styles.css")
def public_styles():
    return send_from_directory(RESOURCE_DIR, "styles.css")


@app.get("/vendor/highcharts/<path:filename>")
@auth_required
def highcharts_asset(filename: str):
    if filename != "highstock.js":
        return jsonify({"error": "Asset not found."}), 404
    return send_from_directory(RESOURCE_DIR / "vendor" / "highcharts", filename)


@app.get("/")
@auth_required
def index():
    return send_from_directory(RESOURCE_DIR, "index.html")


@app.get("/profile")
@auth_required
def profile_page():
    return send_from_directory(RESOURCE_DIR, "profile.html")


@app.get("/api/health")
def health():
    return jsonify({"ok": True, "server": "mt5-backtest", "version": 3})


@app.get("/api/defaults")
@auth_required
def defaults():
    start, end = default_dates()
    return jsonify(
        {
            "from_date": start,
            "to_date": end,
            "sources": SOURCES,
            "source_timeframes": SOURCE_TIMEFRAMES,
            "defaults": {
                "data_source": "MT5",
                "timeframe": "M5",
                "trail_timeframe": "M5",
                "symbol": "GOLD.i#",
                "entry_pattern": "BOTH",
                "range_start": "08:30",
                "range_end": "09:30",
                "session_start": "09:30",
                "entry_cutoff": "19:00",
                "session_end": "21:30",
                "entry_buffer_pct": 0.02,
                "stop_points": 10,
                "stop_points_unit": "POINTS",
                "first_trail_profit": 10,
                "first_trail_profit_unit": "POINTS",
                "first_trail_lock_loss": 0,
                "first_trail_lock_loss_unit": "POINTS",
                "second_trail_profit": 20,
                "second_trail_profit_unit": "POINTS",
            },
        }
    )


@app.get("/api/symbols")
@auth_required
def symbols():
    try:
        source = normalize_source(request.args.get("source", "DELTA"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    if source == "DELTA":
        try:
            return jsonify(delta_symbols())
        except Exception as exc:
            return jsonify({"error": f"Delta symbol list failed: {exc}"}), 400

    if not mt5.initialize():
        return jsonify(FALLBACK_MT5_SYMBOLS)

    try:
        all_symbols = mt5.symbols_get()
    finally:
        mt5.shutdown()

    if not all_symbols:
        return jsonify(FALLBACK_MT5_SYMBOLS)

    priority = ("BTC", "ETH", "XRP", "SOL", "DOGE", "BNB", "ADA", "XAU", "XAG", "US30", "NAS", "SPX")
    names = sorted({symbol.name for symbol in all_symbols} | set(FALLBACK_MT5_SYMBOLS))
    names.sort(key=lambda name: (not any(term in name.upper() for term in priority), name.upper()))
    return jsonify(names)


@app.get("/api/history-range")
@auth_required
def history_range():
    try:
        source = normalize_source(request.args.get("source", "DELTA"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    symbol = request.args.get("symbol", "BTCUSD").strip()
    timeframe = request.args.get("timeframe", "M5").upper()

    default_from, default_to = default_dates()
    requested_from = request.args.get("from_date", default_from)
    requested_to = request.args.get("to_date", default_to)
    try:
        validate_source_timeframe(source, timeframe)
        from_date = parse_date(requested_from)
        to_date = parse_date(requested_to)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    cache_key = (source, symbol, timeframe, requested_from, requested_to)
    cached = _history_cache.get(cache_key)
    if cached and monotonic() - cached[0] < HISTORY_CACHE_SECONDS:
        return jsonify(cached[1])

    if source == "DELTA":
        try:
            result = delta_history_status(symbol, timeframe, from_date, to_date)
        except Exception as exc:
            return jsonify({"error": str(exc)}), 400
        _history_cache[cache_key] = (monotonic(), result)
        return jsonify(result)

    try:
        config = BacktestConfig(
            symbol=symbol,
            from_date=from_date,
            to_date=to_date,
            data_source=source,
            timeframe=timeframe,
        )
        df = fetch_rates(config)
    except Exception as exc:
        if "No MT5 candle data returned" not in str(exc):
            return jsonify({"error": str(exc)}), 400
        result = {"symbol": symbol, "timeframe": timeframe, "available": False}
    else:
        if df.empty:
            result = {"symbol": symbol, "timeframe": timeframe, "available": False}
        else:
            result = {
                "symbol": symbol,
                "timeframe": timeframe,
                "available": True,
                "from_date": min(df["trade_date"]).isoformat(),
                "to_date": max(df["trade_date"]).isoformat(),
                "from_time": df.iloc[0]["time_ist"].isoformat(),
                "to_time": df.iloc[-1]["time_ist"].isoformat(),
            }
    if result["available"]:
        _history_cache[cache_key] = (monotonic(), result)
    return jsonify(result)


@app.get("/api/export/excel")
@auth_required
def export_excel():
    try:
        summary = load_latest_summary(request.args.get("source", "DELTA"))
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        config_rows = [{"field": key, "value": value} for key, value in summary.get("config", {}).items()]
        stats_rows = [{"metric": key, "value": value} for key, value in summary.get("stats", {}).items()]
        trades = summary.get("trades", [])

        add_sheet_rows(workbook, "Config", config_rows)
        add_sheet_rows(workbook, "Stats", stats_rows)
        add_sheet_rows(workbook, "Trades", trades)
        add_sheet_rows(workbook, "Daily Points", grouped_points(trades, "day"))
        add_sheet_rows(workbook, "Monthly Points", grouped_points(trades, "month"))
        add_sheet_rows(workbook, "Yearly Points", grouped_points(trades, "year"))

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        symbol = summary.get("config", {}).get("symbol", "backtest")
        filename = f"{symbol}_backtest_export.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/data/clear")
@auth_required
def clear_all_data():
    if not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    with _saved_data_lock:
        with _optimizer_lock:
            active_job = next(
                (job for job in _optimizer_jobs.values() if job.get("status") in {"queued", "running"}),
                None,
            )
            if active_job:
                return jsonify({"error": "An optimizer scan is running. Wait for it to finish before clearing all data."}), 409
            try:
                clear_saved_data()
                clear_python_caches()
                _optimizer_jobs.clear()
                _history_cache.clear()
                clear_rates_cache()
                clear_market_data_cache()
            except OSError as exc:
                return jsonify({"error": f"Could not clear saved data: {exc}"}), 500
    return jsonify({"message": "All saved results, candle data, and Python caches have been cleared."})


@app.post("/api/backtest")
@auth_required
@serialized_saved_data_write
def run_backtest():
    if not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    data = request.get_json(force=True)
    try:
        source = normalize_source(data.get("data_source", "DELTA"))
        config = BacktestConfig(
            symbol=str(data.get("symbol", "")).strip(),
            from_date=parse_date(data["from_date"]),
            to_date=parse_date(data["to_date"]),
            data_source=source,
            timeframe=str(data.get("timeframe", "M5")).upper(),
            trail_timeframe=str(data.get("trail_timeframe", "M15")).upper(),
            entry_pattern=str(data.get("entry_pattern", "BOTH")).upper(),
            range_start=parse_time(data.get("range_start", "08:30")),
            range_end=parse_time(data.get("range_end", "09:30")),
            session_start=parse_time(data.get("session_start", "09:30")),
            entry_cutoff=parse_time(data.get("entry_cutoff", "18:00")),
            session_end=parse_time(data.get("session_end", "19:30")),
            entry_buffer_pct=float(data.get("entry_buffer_pct", 0.25)) / 100,
            stop_points=float(data.get("stop_points", 400)),
            first_trail_profit=float(data.get("first_trail_profit", 400)),
            first_trail_lock_loss=float(data.get("first_trail_lock_loss", 300)),
            second_trail_profit=float(data.get("second_trail_profit", 700)),
            stop_points_unit=data.get("stop_points_unit", "POINTS"),
            first_trail_profit_unit=data.get("first_trail_profit_unit", "POINTS"),
            first_trail_lock_loss_unit=data.get("first_trail_lock_loss_unit", "POINTS"),
            second_trail_profit_unit=data.get("second_trail_profit_unit", "POINTS"),
        )
        if not config.symbol:
            raise ValueError("Symbol is required.")
        validate_source_timeframe(source, config.timeframe)
        validate_source_timeframe(source, config.trail_timeframe)
        if config.entry_pattern not in OPTIMIZER_ENTRY_PATTERNS:
            raise ValueError("Unsupported entry pattern.")
        if config.range_start >= config.range_end:
            raise ValueError("Range start must be before range end.")
        if config.session_start >= config.session_end:
            raise ValueError("Session start must be before session end.")
        if config.entry_cutoff < config.session_start or config.entry_cutoff > config.session_end:
            raise ValueError("Last entry time must be within the trading session.")

        df = fetch_source_rates(config)
        if not df.empty:
            config.data_from_date = min(df["trade_date"])
            config.data_to_date = max(df["trade_date"])
        if config.trail_timeframe == config.timeframe:
            trail_df = df
        else:
            trail_config = BacktestConfig(
                symbol=config.symbol,
                from_date=config.from_date - timedelta(days=1),
                to_date=config.to_date,
                data_source=config.data_source,
                timeframe=config.trail_timeframe,
            )
            trail_df = fetch_source_rates(trail_config)
        if config.timeframe == "M1":
            execution_df = df
        else:
            execution_config = BacktestConfig(
                symbol=config.symbol,
                from_date=config.from_date,
                to_date=config.to_date,
                data_source=config.data_source,
                timeframe="M1",
            )
            execution_df = fetch_source_rates(execution_config)
        trades = backtest(df, config, trail_df, execution_df)
        summary = build_summary(trades, config)
        write_outputs(summary)
        return jsonify(summary)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/optimizer/start")
@auth_required
def start_optimizer():
    if not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    try:
        payload = optimizer_payload(request.get_json(force=True))
        common = payload["common"]
        with _optimizer_lock:
            active_job = next(
                (
                    job
                    for job in _optimizer_jobs.values()
                    if job.get("status") in {"queued", "running"}
                    and job.get("source") == common["data_source"]
                    and job.get("symbol") == common["symbol"]
                ),
                None,
            )
        if active_job:
            return jsonify(
                {
                    "job_id": active_job["scan_id"],
                    "scan_id": active_job["scan_id"],
                    "source": common["data_source"],
                    "symbol": common["symbol"],
                    "total": active_job["total"],
                    "save_full_csv": active_job.get("save_full_csv", False),
                    "reused": True,
                }
            )
        job_id = f"{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}_{secrets.token_hex(4)}"
        created_at = datetime.now(timezone.utc).isoformat()
        job = {
            "scan_id": job_id,
            "source": common["data_source"],
            "symbol": common["symbol"],
            "status": "queued",
            "message": "Scan queued...",
            "tested": 0,
            "total": payload["total"],
            "progress_pct": 0.0,
            "target_win_rate": payload["target_win_rate"],
            "minimum_trades": payload["minimum_trades"],
            "result_sort": payload["result_sort"],
            "save_full_csv": payload["save_full_csv"],
            "created_at": created_at,
            "updated_at": created_at,
            "result_count": 0,
            "results_truncated": False,
            "results": [],
        }
        with _optimizer_lock:
            _optimizer_jobs[job_id] = job
        save_optimizer_output(job, payload, [])
        future = _optimizer_executor.submit(execute_optimizer, job_id, payload)
        future.add_done_callback(lambda item, scan_id=job_id: mark_optimizer_submit_error(scan_id, item))
        return jsonify(
            {
                "job_id": job_id,
                "scan_id": job_id,
                "source": common["data_source"],
                "symbol": common["symbol"],
                "total": payload["total"],
                "save_full_csv": payload["save_full_csv"],
            }
        )
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/optimizer/jobs/<job_id>")
@auth_required
def optimizer_job(job_id: str):
    with _optimizer_lock:
        job = _optimizer_jobs.get(job_id)
        result = dict(job) if job else None
    if result is None:
        return jsonify({"error": "Optimizer scan not found."}), 404
    return jsonify(result)


@app.get("/api/optimizer/scans")
@auth_required
def optimizer_scans():
    try:
        source = normalize_source(request.args.get("source", "DELTA"))
        symbol = request.args.get("symbol", "").strip()
        if not symbol:
            raise ValueError("Symbol is required.")
        records = saved_optimizer_scans(source, symbol, include_results=False)
        summary = [public_saved_optimizer_scan(row) for row in records[:50]]
        return jsonify(summary)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/optimizer/scans/<scan_id>")
@auth_required
def saved_optimizer_scan(scan_id: str):
    try:
        source = normalize_source(request.args.get("source", "DELTA"))
        symbol = request.args.get("symbol", "").strip()
        if not symbol:
            raise ValueError("Symbol is required.")
        result = load_saved_optimizer_scan(source, symbol, scan_id)
        if result is None:
            return jsonify({"error": "Optimizer scan not found."}), 404
        with _optimizer_lock:
            active = scan_id in _optimizer_jobs
        if result.get("status") in {"queued", "running"} and not active:
            result["status"] = "interrupted"
            result["message"] = "Scan was interrupted. Saved partial results are available."
        return jsonify(result)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 409
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/optimizer/latest")
@auth_required
def latest_optimizer():
    try:
        source = normalize_source(request.args.get("source", "DELTA"))
        symbol = request.args.get("symbol", "").strip()
        if not symbol:
            raise ValueError("Symbol is required.")
        scans = saved_optimizer_scans(source, symbol)
        if not scans:
            return jsonify({"error": "No optimizer scan result found."}), 404
        return jsonify(scans[0])
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/optimizer/export")
@auth_required
def export_optimizer():
    try:
        source = normalize_source(request.args.get("source", "DELTA"))
        symbol = request.args.get("symbol", "").strip()
        if not symbol:
            raise ValueError("Symbol is required.")
        scan_id = request.args.get("scan_id", "").strip()
        if not scan_id:
            scans = saved_optimizer_scans(source, symbol)
            if not scans:
                return jsonify({"error": "No optimizer scan result found."}), 404
            scan_id = scans[0]["scan_id"]
        path = optimizer_csv_path(source, symbol, scan_id)
        if not path.exists():
            scan = load_saved_optimizer_scan(source, symbol, scan_id)
            if scan is None:
                return jsonify({"error": "No optimizer scan result found."}), 404
            rows = scan.get("results") or []
            if not rows:
                return jsonify({"error": "No optimizer scan result found."}), 404
            text_buffer = StringIO()
            writer = csv.DictWriter(text_buffer, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
            bytes_buffer = BytesIO(text_buffer.getvalue().encode("utf-8"))
            filename = f"{safe_optimizer_part(symbol).lower()}_{source.lower()}_{scan_id}_optimizer.csv"
            return send_file(bytes_buffer, mimetype="text/csv", as_attachment=True, download_name=filename)
        filename = f"{safe_optimizer_part(symbol).lower()}_{source.lower()}_{scan_id}_optimizer.csv"
        return send_file(path, as_attachment=True, download_name=filename)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 409
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/candles", methods=["GET", "POST", "OPTIONS"])
@auth_required
def candles():
    if request.method == "OPTIONS":
        return jsonify({"ok": True})
    if request.method == "POST" and not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    data = request.args.to_dict() if request.method == "GET" else request.get_json(force=True)
    try:
        source = normalize_source(data.get("data_source", "DELTA"))
        from_date = parse_date(data.get("from_date") or data["trade_date"])
        to_date = parse_date(data.get("to_date") or data.get("trade_date") or data["from_date"])
        config = BacktestConfig(
            symbol=str(data.get("symbol", "")).strip(),
            from_date=from_date,
            to_date=to_date,
            data_source=source,
            timeframe=str(data.get("timeframe", "M5")).upper(),
        )
        if not config.symbol:
            raise ValueError("Symbol is required.")
        validate_source_timeframe(source, config.timeframe)
        df = fetch_source_rates(config)
        rows = []
        for _, row in df.iterrows():
            rows.append(
                {
                    "time": int(row["time_ist"].timestamp()),
                    "open": round(float(row["open"]), 2),
                    "high": round(float(row["high"]), 2),
                    "low": round(float(row["low"]), 2),
                    "close": round(float(row["close"]), 2),
                }
            )
        return jsonify({"candles": rows, "timezone": "Asia/Kolkata"})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/session")
@auth_required
def current_session():
    return jsonify({"username": session["username"], "csrf_token": csrf_token()})


@app.route("/api/profile", methods=["GET", "POST"])
@auth_required
def profile_details():
    if request.method == "POST" and not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    try:
        if request.method == "POST":
            return jsonify(save_profile(request.get_json(force=True)))
        return jsonify(load_profile())
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/profile/password")
@auth_required
def update_profile_password():
    if not csrf_is_valid():
        return jsonify({"error": "Invalid security token."}), 400
    try:
        change_password(request.get_json(force=True))
        return jsonify({"message": "Password updated successfully."})
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/report_data.js")
@auth_required
def latest_report_javascript():
    path = REPORT_DATA_FILE
    if path.is_file():
        return send_from_directory(path.resolve().parent, path.name)
    return app.response_class("window.BACKTEST_REPORT = null;\n", mimetype="application/javascript")


@app.get("/<path:path>")
@auth_required
def static_files(path: str):
    if path == "algo" or path in {"algo.css", "algo.js"} or path.startswith("api/algo"):
        if path.startswith("api/"):
            return jsonify({"error": "Algo runs as a separate app."}), 404
        return ("Algo runs as a separate app.", 404)
    return send_from_directory(RESOURCE_DIR, "index.html")


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
